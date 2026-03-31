#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV 数据导入工具
支持 SQLite / MySQL / Oracle
"""

import sys
import os
import csv
import json
import logging
import threading
from collections import Counter
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from i18n import _, set_lang, current_lang, available_langs, LANG_LABELS

# ── 可选依赖 ──────────────────────────────────────────────────────────────────
try:
    import mysql.connector
    HAS_MYSQL = True
except ImportError:
    HAS_MYSQL = False

try:
    import oracledb
    HAS_ORACLE = True
except ImportError:
    HAS_ORACLE = False

try:
    import xlsxwriter as _xlsxwriter   # noqa: F401
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

try:
    import openpyxl as _openpyxl       # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Oracle Thick 模式只初始化一次
_oracle_thick_initialized = False


def _auto_detect_oracle_client() -> str:
    """扫描 APP_DIR 下的 instantclient_* 目录，返回第一个找到的路径。"""
    import glob
    pattern = os.path.join(APP_DIR, "instantclient*")
    matches = sorted(glob.glob(pattern))
    return matches[0] if matches else ""


def _resolve_lib_dir(lib_dir: str) -> str:
    """将相对路径转为绝对路径（相对于 APP_DIR）。"""
    if not lib_dir:
        return ""
    p = os.path.expandvars(lib_dir)          # 展开 %VAR%
    if not os.path.isabs(p):
        p = os.path.join(APP_DIR, p)         # 相对路径 → 绝对路径
    return os.path.normpath(p)


def _init_oracle_client(lib_dir: str = "") -> None:
    global _oracle_thick_initialized
    if _oracle_thick_initialized:
        return
    resolved = _resolve_lib_dir(lib_dir) or _auto_detect_oracle_client()
    if not resolved:
        raise FileNotFoundError(
            _("oracle.err.no_client", app_dir=APP_DIR)
        )
    if not os.path.isdir(resolved):
        raise FileNotFoundError(
            _("oracle.err.bad_dir", resolved=resolved, app_dir=APP_DIR)
        )
    # Python 3.8+ / PyInstaller exe 必须显式注册 DLL 搜索目录
    # 否则 Windows 找不到 oci.dll，导致 DPI-1072
    if hasattr(os, "add_dll_directory"):
        os.add_dll_directory(resolved)

    # 同时加入 PATH（兼容旧版 Windows 加载方式）
    os.environ["PATH"] = resolved + os.pathsep + os.environ.get("PATH", "")

    oracledb.init_oracle_client(lib_dir=resolved)
    print(f"[Oracle] Thick mode, Instant Client: {resolved}", flush=True)
    print(f"[Oracle] oracledb version: {oracledb.__version__}", flush=True)
    _oracle_thick_initialized = True

try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

# ── 全局路径 ──────────────────────────────────────────────────────────────────
# PyInstaller 单文件 exe 运行时 sys.argv[0] 指向 exe 本身，取其所在目录
# 开发模式下取 __file__ 所在目录
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE    = os.path.join(APP_DIR, "db_config.json")
UI_STATE_FILE  = os.path.join(APP_DIR, "ui_state.json")
LOG_DIR        = os.path.join(APP_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)


def _load_ui_state() -> dict:
    try:
        if os.path.exists(UI_STATE_FILE):
            with open(UI_STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _save_ui_state(state: dict):
    try:
        existing = _load_ui_state()
        existing.update(state)
        with open(UI_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# 启动时恢复上次选择的语言
_saved_lang = _load_ui_state().get("lang", "zh_CN")
set_lang(_saved_lang)


# ─────────────────────────────────────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────────────────────────────────────

from file_utils import (
    detect_encoding,
    detect_encoding_from_bytes,
    count_lines      as _fu_count_lines,
    read_head        as _fu_read_head,
    read_tail        as _fu_read_tail,
    read_line_at     as _fu_read_line_at,
    read_last_nth_line as _fu_read_last_nth,
    split_file       as _fu_split_file,
)


def parse_ignore_ranges(text: str) -> set:
    """
    解析忽略行区间文本，返回行号集合（1-based）。
    格式示例: '1,10-20,54,55-80'
    """
    result = set()
    for part in text.split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            lo, _, hi = part.partition('-')
            lo, hi = lo.strip(), hi.strip()
            if lo.isdigit() and hi.isdigit():
                result.update(range(int(lo), int(hi) + 1))
        elif part.isdigit():
            result.add(int(part))
    return result


def normalize_ranges(row_set: set) -> str:
    """
    将行号集合归一化为紧凑区间文本（排序、去重、合并相邻）。
    {1,10,11,12,20} → '1,10-12,20'
    """
    if not row_set:
        return ''
    nums = sorted(row_set)
    parts = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
        else:
            parts.append(str(start) if start == prev else f'{start}-{prev}')
            start = prev = n
    parts.append(str(start) if start == prev else f'{start}-{prev}')
    return ','.join(parts)


def read_rows(path: str, encoding: str, delimiter: str, quotechar: str):
    """
    Generator: 逐行解析文件，返回字段列表。
    单字符分隔符：使用 csv.reader（支持引用字符、跨行字段）。
    多字符分隔符：使用 str.split()（quotechar 无效）。
    """
    if len(delimiter) == 1:
        with open(path, "r", encoding=encoding, errors="replace", newline="") as fh:
            yield from csv.reader(fh, delimiter=delimiter, quotechar=quotechar or '"')
    else:
        with open(path, "r", encoding=encoding, errors="replace") as fh:
            for line in fh:
                yield line.rstrip("\n\r").split(delimiter)


def make_logger(log_path: str) -> logging.Logger:
    logger = logging.getLogger(f"csv_importer_{datetime.now().strftime('%H%M%S')}")
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(fh)
    return logger


# ─────────────────────────────────────────────────────────────────────────────
# 数据库连接配置管理
# ─────────────────────────────────────────────────────────────────────────────

class DbConfigManager:
    def __init__(self):
        self.configs: dict = {}
        self._load()

    def _load(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    self.configs = json.load(f)
            except Exception:
                self.configs = {}

    def save(self, name: str, cfg: dict):
        self.configs[name] = cfg
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.configs, f, ensure_ascii=False, indent=2)

    def delete(self, name: str):
        self.configs.pop(name, None)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.configs, f, ensure_ascii=False, indent=2)

    def names(self) -> list:
        return list(self.configs.keys())

    def get(self, name: str) -> dict:
        return self.configs.get(name, {})


# ─────────────────────────────────────────────────────────────────────────────
# 数据库连接抽象
# ─────────────────────────────────────────────────────────────────────────────

class DBConnection:
    """封装 SQLite / MySQL / Oracle 的差异。"""

    def __init__(self, cfg: dict):
        self.cfg = cfg
        self.db_type: str = cfg.get("type", "sqlite")
        self.conn = None

    # ── 连接 / 断开 ──────────────────────────────────────────────────────────
    def connect(self):
        t = self.db_type
        if t == "sqlite":
            import sqlite3
            path = self.cfg.get("path", "")
            if not path:
                raise ValueError(_("db.err.sqlite_no_path"))
            # 相对路径 → 相对于程序所在目录（双击 EXE 时工作目录不确定）
            if not os.path.isabs(path):
                path = os.path.join(APP_DIR, path)
            os.makedirs(os.path.dirname(path), exist_ok=True)
            self.conn = sqlite3.connect(path)
        elif t == "mysql":
            if not HAS_MYSQL:
                raise ImportError(_("db.err.no_mysql"))
            self.conn = mysql.connector.connect(
                host=self.cfg.get("host", "localhost"),
                port=int(self.cfg.get("port", 3306)),
                user=self.cfg.get("user", ""),
                password=self.cfg.get("password", ""),
                database=self.cfg.get("database", ""),
                connect_timeout=10,
                charset="utf8mb4",
            )
        elif t == "oracle":
            if not HAS_ORACLE:
                raise ImportError(_("db.err.no_oracle"))
            lib_dir = self.cfg.get("lib_dir", "") or ""
            _init_oracle_client(lib_dir)
            dsn = oracledb.makedsn(
                self.cfg.get("host", ""),
                int(self.cfg.get("port", 1521)),
                service_name=self.cfg.get("service", ""),
            )
            self.conn = oracledb.connect(
                user=self.cfg.get("user", ""),
                password=self.cfg.get("password", ""),
                dsn=dsn,
            )
        else:
            raise ValueError(_("db.err.unsupported", t=t))
        return self

    def close(self):
        if self.conn:
            try:
                self.conn.close()
            except Exception:
                pass
            self.conn = None

    def cursor(self):
        return self.conn.cursor()

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    # ── SQL 方言辅助 ──────────────────────────────────────────────────────────
    def quote(self, name: str) -> str:
        if self.db_type == "mysql":
            return f"`{name}`"
        elif self.db_type == "oracle":
            return f'"{name.upper()}"'
        return f'"{name}"'

    def placeholders(self, columns: list) -> str:
        """返回 INSERT 的占位符字符串。"""
        if self.db_type == "mysql":
            return ", ".join(["%s"] * len(columns))
        elif self.db_type == "oracle":
            # Oracle 位置绑定变量 :1, :2, ... 避免列名含中文等非 ASCII 字符时命名冲突
            return ", ".join([f":{i + 1}" for i in range(len(columns))])
        return ", ".join(["?"] * len(columns))

    def make_row(self, values: list, columns: list):
        """返回一行数据：Oracle 空字符串转 None，其余数据库保持 tuple。"""
        if self.db_type == "oracle":
            return tuple(None if v == "" else v for v in values)
        return tuple(values)

    # ── 元数据查询 ────────────────────────────────────────────────────────────
    def table_exists(self, table: str) -> bool:
        cur = self.cursor()
        t = self.db_type
        if t == "sqlite":
            cur.execute('SELECT 1 FROM sqlite_master WHERE type="table" AND name=?', (table,))
        elif t == "mysql":
            cur.execute("SHOW TABLES LIKE %s", (table,))
        elif t == "oracle":
            cur.execute(
                "SELECT 1 FROM user_tables WHERE table_name=:1",
                (table.upper(),),
            )
        return cur.fetchone() is not None

    def get_table_columns(self, table: str) -> list:
        cur = self.cursor()
        t = self.db_type
        if t == "sqlite":
            cur.execute(f'PRAGMA table_info("{table}")')
            return [row[1] for row in cur.fetchall()]
        elif t == "mysql":
            cur.execute(f"DESCRIBE `{table}`")
            return [row[0] for row in cur.fetchall()]
        elif t == "oracle":
            cur.execute(
                "SELECT column_name FROM user_tab_columns WHERE table_name=:1 ORDER BY column_id",
                (table.upper(),),
            )
            return [row[0] for row in cur.fetchall()]
        return []

    def create_table(self, table: str, columns: list):
        col_defs = ", ".join([f"{self.quote(c)} VARCHAR2(4000)" if self.db_type == "oracle"
                               else f"{self.quote(c)} TEXT"
                               for c in columns])
        sql = f"CREATE TABLE {self.quote(table)} ({col_defs})"
        self.cursor().execute(sql)
        self.commit()

    def truncate_table(self, table: str):
        self.cursor().execute(f"DELETE FROM {self.quote(table)}")
        self.commit()


# ─────────────────────────────────────────────────────────────────────────────
# 数据预览表格控件（Canvas-based，支持奇偶列变色、可编辑表头、自动列宽）
# ─────────────────────────────────────────────────────────────────────────────

class _PreviewGrid(tk.Frame):
    """
    Canvas-based 数据预览表格，替代 ttk.Treeview，支持：
      - 奇偶列不同背景色
      - 双击表头列名可编辑
      - 按内容自动计算初始列宽
      - 拖拽列边缘调整宽度
      - 双向滚动条（横向同步表头）
    """

    # 颜色方案
    COL_COLORS = ("#EEF4FF", "#F5FAF0")   # 奇/偶列单元格背景
    HDR_COLORS = ("#C5D9F1", "#D8E4BC")   # 奇/偶列表头背景
    HDR_FG     = "#1A1A2E"                # 表头文字颜色
    CELL_FG    = "#222222"                # 单元格文字颜色
    SEP_COLOR  = "#AAAAAA"               # 列分隔线颜色
    ROW_LINE   = "#DDDDDD"               # 行分隔线颜色

    HDR_HEIGHT = 26
    ROW_HEIGHT = 20
    PAD_X      = 5
    FONT       = ("Consolas", 9)
    MIN_COL_W  = 55
    MAX_COL_W  = 320
    PX_PER_CHAR = 7   # Consolas 9pt 每字符大约宽度(像素)

    def __init__(self, master, **kw):
        super().__init__(master, **kw)

        # ── 表头 Canvas（高度固定，只横向滚动）──────────────────────────────
        self._hdr_canvas = tk.Canvas(
            self, height=self.HDR_HEIGHT, bg=self.HDR_COLORS[0],
            highlightthickness=0)
        self._hdr_canvas.pack(side=tk.TOP, fill=tk.X)

        # ── 数据区：Canvas + 双向滚动条 ──────────────────────────────────────
        data_frame = tk.Frame(self)
        data_frame.pack(fill=tk.BOTH, expand=True)

        self._vsb = ttk.Scrollbar(data_frame, orient=tk.VERTICAL)
        self._hsb = ttk.Scrollbar(data_frame, orient=tk.HORIZONTAL)
        self._data_canvas = tk.Canvas(
            data_frame, bg="white", highlightthickness=0,
            yscrollcommand=self._vsb.set,
            xscrollcommand=self._on_xscroll)
        self._vsb.configure(command=self._data_canvas.yview)
        self._hsb.configure(command=self._on_hscroll)

        self._vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._data_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ── 内部状态 ─────────────────────────────────────────────────────────
        self._columns:    list  = []   # 原始列名
        self._col_names:  list  = []   # 用户编辑后的列名
        self._col_widths: list  = []   # 各列像素宽度
        self._rows:       list  = []   # 当前页数据行

        self._resize_col:     int  = -1    # 正在拖拽的列索引
        self._resize_start_x: float = 0
        self._resize_start_w: int   = 0

        self._edit_entry = None   # 活跃的表头编辑 Entry

        # ── 事件绑定 ─────────────────────────────────────────────────────────
        self._hdr_canvas.bind("<Motion>",          self._on_hdr_motion)
        self._hdr_canvas.bind("<ButtonPress-1>",   self._on_hdr_press)
        self._hdr_canvas.bind("<B1-Motion>",       self._on_hdr_drag)
        self._hdr_canvas.bind("<ButtonRelease-1>", self._on_hdr_release)
        self._hdr_canvas.bind("<Double-Button-1>", self._on_hdr_double)

        self._data_canvas.bind("<MouseWheel>", self._on_mousewheel)
        self._data_canvas.bind("<Configure>",  lambda e: self._draw_data())

    # ── 公共接口 ──────────────────────────────────────────────────────────────

    def set_data(self, columns: list, rows: list, saved_widths: dict = None):
        """
        加载新数据并重绘。
          columns      : 原始列名列表
          rows         : 数据行列表，每行是字段值列表
          saved_widths : {列名: 宽度} 上次保存的列宽（可选）
        """
        self._columns   = list(columns)
        self._col_names = list(columns)
        self._rows      = rows

        # 计算初始列宽
        self._col_widths = []
        for i, col in enumerate(columns):
            if saved_widths and col in saved_widths:
                w = saved_widths[col]
            else:
                max_data = 0
                for row in rows[:80]:
                    if i < len(row):
                        cell = str(row[i])
                        # 中文字符约占 2 个英文字符宽
                        cw = sum(2 if ord(ch) > 127 else 1 for ch in cell)
                        max_data = max(max_data, cw)
                col_cw = sum(2 if ord(ch) > 127 else 1 for ch in col)
                max_len = max(col_cw, max_data)
                w = max_len * self.PX_PER_CHAR + self.PAD_X * 2
                w = max(self.MIN_COL_W, min(self.MAX_COL_W, w))
            self._col_widths.append(int(w))

        self._draw_header()
        self._draw_data()

    def get_col_widths(self) -> dict:
        """返回 {原始列名: 当前宽度} 供外部保存。"""
        return {col: w for col, w in zip(self._columns, self._col_widths)}

    def get_col_names(self) -> list:
        """返回用户编辑后的列名列表。"""
        return list(self._col_names)

    # ── 滚动联动 ──────────────────────────────────────────────────────────────

    def _on_xscroll(self, *args):
        """数据 Canvas 横向滚动 → 同步表头。"""
        self._hsb.set(*args)
        self._hdr_canvas.xview_moveto(args[0])

    def _on_hscroll(self, *args):
        """滚动条拖动 → 同步两个 Canvas。"""
        self._data_canvas.xview(*args)
        self._hdr_canvas.xview(*args)

    def _on_mousewheel(self, event):
        self._data_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ── 绘制 ──────────────────────────────────────────────────────────────────

    def _total_width(self) -> int:
        return sum(self._col_widths) if self._col_widths else 0

    def _col_x(self, idx: int) -> int:
        return sum(self._col_widths[:idx])

    def _draw_header(self):
        c = self._hdr_canvas
        c.delete("all")
        if not self._col_names:
            return
        total_w = self._total_width()
        c.configure(scrollregion=(0, 0, total_w, self.HDR_HEIGHT))

        x = 0
        for i, name in enumerate(self._col_names):
            w = self._col_widths[i]
            bg = self.HDR_COLORS[i % 2]
            c.create_rectangle(x, 0, x + w - 1, self.HDR_HEIGHT - 1,
                                fill=bg, outline=self.SEP_COLOR)
            # 截断超长表头文字
            max_chars = max(1, (w - self.PAD_X * 2) // self.PX_PER_CHAR)
            display = name if len(name) <= max_chars else name[:max_chars - 1] + "…"
            c.create_text(x + self.PAD_X, self.HDR_HEIGHT // 2,
                          text=display, anchor=tk.W,
                          font=self.FONT, fill=self.HDR_FG)
            x += w

    def _draw_data(self):
        c = self._data_canvas
        c.delete("all")
        if not self._columns or not self._rows:
            return

        total_w = self._total_width()
        total_h = len(self._rows) * self.ROW_HEIGHT
        c.configure(scrollregion=(0, 0, total_w, max(total_h, 1)))

        # 只绘制可见行（优化大页面性能）
        try:
            canvas_h = c.winfo_height() or 600
        except Exception:
            canvas_h = 600
        y_top, y_bot = 0, canvas_h + 200   # 稍微多画一些避免闪烁
        try:
            frac_top = float(c.yview()[0])
            y_top = int(frac_top * total_h) - self.ROW_HEIGHT
        except Exception:
            y_top = 0

        first_row = max(0, y_top // self.ROW_HEIGHT)
        last_row  = min(len(self._rows), first_row + canvas_h // self.ROW_HEIGHT + 4)

        for row_idx in range(first_row, last_row):
            row = self._rows[row_idx]
            y = row_idx * self.ROW_HEIGHT
            x = 0
            for col_idx in range(len(self._columns)):
                w = self._col_widths[col_idx]
                bg = self.COL_COLORS[col_idx % 2]
                c.create_rectangle(x, y, x + w - 1, y + self.ROW_HEIGHT - 1,
                                   fill=bg, outline=self.ROW_LINE)
                val = str(row[col_idx]) if col_idx < len(row) else ""
                # 截断超长内容
                max_chars = max(1, (w - self.PAD_X * 2) // self.PX_PER_CHAR)
                display = val if len(val) <= max_chars else val[:max_chars - 1] + "…"
                c.create_text(x + self.PAD_X, y + self.ROW_HEIGHT // 2,
                              text=display, anchor=tk.W,
                              font=self.FONT, fill=self.CELL_FG)
                x += w

    # ── 表头交互 ──────────────────────────────────────────────────────────────

    def _sep_col_at(self, canvas_x: float) -> int:
        """返回 canvas_x 处对应的列分隔线右侧列索引，找不到返回 -1。"""
        x = 0
        for i, w in enumerate(self._col_widths):
            x += w
            if abs(canvas_x - x) <= 5:
                return i
        return -1

    def _col_at(self, canvas_x: float) -> int:
        """返回 canvas_x 所在列索引，找不到返回 -1。"""
        x = 0
        for i, w in enumerate(self._col_widths):
            if x <= canvas_x < x + w:
                return i
            x += w
        return -1

    def _on_hdr_motion(self, event):
        cx = self._hdr_canvas.canvasx(event.x)
        if self._sep_col_at(cx) >= 0:
            self._hdr_canvas.configure(cursor="sb_h_double_arrow")
        else:
            self._hdr_canvas.configure(cursor="")

    def _on_hdr_press(self, event):
        cx = self._hdr_canvas.canvasx(event.x)
        sep = self._sep_col_at(cx)
        if sep >= 0:
            self._resize_col     = sep
            self._resize_start_x = cx
            self._resize_start_w = self._col_widths[sep]

    def _on_hdr_drag(self, event):
        if self._resize_col < 0:
            return
        cx = self._hdr_canvas.canvasx(event.x)
        delta = cx - self._resize_start_x
        new_w = max(self.MIN_COL_W, int(self._resize_start_w + delta))
        self._col_widths[self._resize_col] = new_w
        self._draw_header()
        self._draw_data()

    def _on_hdr_release(self, event):
        self._resize_col = -1

    def _on_hdr_double(self, event):
        """双击表头单元格 → 弹出内联 Entry 编辑列名。"""
        cx = self._hdr_canvas.canvasx(event.x)
        if self._sep_col_at(cx) >= 0:
            return   # 不在列边缘触发
        col_idx = self._col_at(cx)
        if col_idx < 0:
            return
        self._commit_edit()   # 关闭前一个编辑框
        self._start_rename(col_idx)

    def _start_rename(self, col_idx: int):
        col_x = self._col_x(col_idx)
        col_w = self._col_widths[col_idx]

        # Canvas 坐标 → 控件像素坐标（需减去横向滚动偏移）
        try:
            frac = float(self._hdr_canvas.xview()[0])
            total_w = self._total_width()
            scroll_px = int(frac * total_w)
        except Exception:
            scroll_px = 0
        widget_x = col_x - scroll_px
        widget_y = 2

        var = tk.StringVar(value=self._col_names[col_idx])
        entry = tk.Entry(self._hdr_canvas, textvariable=var,
                         font=self.FONT, relief="solid", bd=1)
        entry.place(x=widget_x, y=widget_y,
                    width=max(col_w - 4, 20), height=self.HDR_HEIGHT - 4)
        entry.focus_set()
        entry.select_range(0, tk.END)
        self._edit_entry = (entry, col_idx, var)

        entry.bind("<Return>",   lambda e: self._commit_edit())
        entry.bind("<Escape>",   lambda e: self._cancel_edit())
        entry.bind("<FocusOut>", lambda e: self._commit_edit())

    def _commit_edit(self):
        if self._edit_entry is None:
            return
        entry, col_idx, var = self._edit_entry
        new_name = var.get().strip()
        if new_name:
            self._col_names[col_idx] = new_name
        try:
            entry.destroy()
        except Exception:
            pass
        self._edit_entry = None
        self._draw_header()

    def _cancel_edit(self):
        if self._edit_entry is None:
            return
        entry, *_ = self._edit_entry
        try:
            entry.destroy()
        except Exception:
            pass
        self._edit_entry = None
        self._draw_header()


# ─────────────────────────────────────────────────────────────────────────────
# 主应用
# ─────────────────────────────────────────────────────────────────────────────

class CSVImporterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(_("app.title"))
        self.geometry("960x700")
        self.minsize(800, 600)

        self.cfg_mgr = DbConfigManager()
        self.csv_encoding_var = tk.StringVar(value="")
        self.ignored_error_types: set = set()
        self.validation_result: dict | None = None
        self._stop_import = threading.Event()  # 停止导入信号
        self._fw_positions_var = tk.StringVar(value="")   # 固定宽度建议分割位置
        self._val_mode_var = tk.StringVar(value="delimiter")  # 校验模式
        # 忽略行数（确认后生效）
        self._skip_head: int = 0
        self._skip_tail: int = 0
        self._skip_middle: set = set()   # 中间行号集合（1-based）
        # 标题行
        self.has_header_var = tk.BooleanVar(value=False)
        self.header_row_var = tk.StringVar(value="")
        # 导出 Tab 状态
        self._export_source_var = tk.StringVar(value="file")
        self._export_format_var = tk.StringVar(value="csv")
        self._export_split_var = tk.StringVar(value="all")
        self._export_include_header_var = tk.BooleanVar(value=True)
        self._export_records_per_file_var = tk.StringVar(value="10000")
        self._stop_export = threading.Event()
        # Excel 引擎：优先 xlsxwriter，fallback openpyxl
        default_engine = "xlsxwriter" if HAS_XLSXWRITER else ("openpyxl" if HAS_OPENPYXL else "")
        self._excel_engine_var = tk.StringVar(value=default_engine)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_path = os.path.join(LOG_DIR, f"import_{ts}.log")
        self.logger = make_logger(self.log_path)

        self._build_ui()
        self._load_last_saved_config()

    # ── UI 构建 ───────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        self._build_db_tab()
        self._build_csv_tab()
        self._build_preview_tab()
        self._build_validate_tab()
        self._build_import_tab()
        self._build_export_tab()
        self._build_info_tab()

    # ══════════════════════════════════════════════════════════════════════════
    # Tab 1 — 数据库连接
    # ══════════════════════════════════════════════════════════════════════════
    def _build_db_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text=_("tab.db"))

        # 已保存连接
        saved = ttk.LabelFrame(f, text=_("db.saved.label"))
        saved.pack(fill=tk.X, padx=12, pady=(12, 4))
        self.saved_name_var = tk.StringVar()
        self.saved_combo = ttk.Combobox(saved, textvariable=self.saved_name_var, width=30, state="readonly")
        self.saved_combo.pack(side=tk.LEFT, padx=6, pady=6)
        ttk.Button(saved, text=_("db.btn.load"), command=self._load_db_config).pack(side=tk.LEFT, padx=3)
        ttk.Button(saved, text=_("db.btn.delete"), command=self._delete_db_config).pack(side=tk.LEFT, padx=3)

        # 连接字段
        fields_frame = ttk.LabelFrame(f, text=_("db.fields.label"))
        fields_frame.pack(fill=tk.X, padx=12, pady=4)

        ttk.Label(fields_frame, text=_("db.type.label")).grid(row=0, column=0, sticky=tk.W, padx=8, pady=5)
        self.db_type_var = tk.StringVar(value="sqlite")
        db_type_cb = ttk.Combobox(fields_frame, textvariable=self.db_type_var,
                                   values=["sqlite", "mysql", "oracle"], width=10, state="readonly")
        db_type_cb.grid(row=0, column=1, sticky=tk.W, padx=6, pady=5)
        db_type_cb.bind("<<ComboboxSelected>>", self._on_db_type_change)

        self._db_vars: dict[str, tk.StringVar] = {}
        self._db_entries: dict[str, ttk.Widget] = {}

        def add_row(row, label, key, default="", show=""):
            ttk.Label(fields_frame, text=label).grid(row=row, column=0, sticky=tk.W, padx=8, pady=4)
            var = tk.StringVar(value=default)
            self._db_vars[key] = var
            e = ttk.Entry(fields_frame, textvariable=var, width=38, show=show)
            e.grid(row=row, column=1, sticky=tk.W, padx=6, pady=4)
            self._db_entries[key] = e
            return e

        self._db_entries["conn_name_lbl"] = ttk.Label(fields_frame, text=_("db.conn_name.label"))
        self._db_entries["conn_name_lbl"].grid(row=1, column=0, sticky=tk.W, padx=8, pady=4)
        self._db_vars["conn_name"] = tk.StringVar(value="")
        e_name = ttk.Entry(fields_frame, textvariable=self._db_vars["conn_name"], width=38)
        e_name.grid(row=1, column=1, sticky=tk.W, padx=6, pady=4)
        self._db_entries["conn_name"] = e_name

        add_row(2, _("db.host.label"), "host", "localhost")
        ttk.Button(fields_frame, text=_("common.btn.browse"), command=self._browse_sqlite_file).grid(
            row=2, column=2, padx=4)

        add_row(3, _("db.port.label"), "port", "3306")
        add_row(4, _("db.user.label"), "user", "")
        add_row(5, _("db.password.label"), "password", "", show="*")
        add_row(6, _("db.database.label"), "database", "")

        # Oracle Instant Client 路径（仅 Oracle Thick 模式需要）
        # 自动检测项目目录下的 instantclient_* 文件夹作为默认值
        _detected = _auto_detect_oracle_client()
        self._db_vars["lib_dir"] = tk.StringVar(value=_detected)
        self._oracle_row = ttk.Frame(fields_frame)
        self._oracle_row.grid(row=7, column=0, columnspan=3, sticky=tk.EW, padx=4, pady=2)
        ttk.Label(self._oracle_row, text=_("db.lib_dir.label")).pack(side=tk.LEFT, padx=4)
        ttk.Entry(self._oracle_row, textvariable=self._db_vars["lib_dir"], width=30).pack(
            side=tk.LEFT, padx=4)
        ttk.Button(self._oracle_row, text=_("common.btn.browse"),
                   command=lambda: self._db_vars["lib_dir"].set(
                       filedialog.askdirectory(title=_("db.browse.oracle.title")) or
                       self._db_vars["lib_dir"].get()
                   )).pack(side=tk.LEFT, padx=2)
        ttk.Label(self._oracle_row,
                  text=_("db.lib_dir.hint"),
                  foreground="gray").pack(side=tk.LEFT, padx=6)

        self._sqlite_hint = ttk.Label(fields_frame,
                                       text=_("db.sqlite.hint"),
                                       foreground="gray")
        self._sqlite_hint.grid(row=8, column=0, columnspan=3, sticky=tk.W, padx=8, pady=2)

        # 操作按钮
        btn_row = ttk.Frame(f)
        btn_row.pack(fill=tk.X, padx=12, pady=8)
        ttk.Button(btn_row, text=_("db.btn.save"), command=self._save_db_config).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_row, text=_("db.btn.test"), command=self._test_connection).pack(side=tk.LEFT, padx=4)
        self.conn_status_var = tk.StringVar(value="")
        ttk.Label(btn_row, textvariable=self.conn_status_var).pack(side=tk.LEFT, padx=10)

        self._refresh_saved_combo()
        self._on_db_type_change()

    def _on_db_type_change(self, _event=None):
        db_type = self.db_type_var.get()
        is_sqlite = db_type == "sqlite"
        is_oracle = db_type == "oracle"
        for key in ("port", "user", "password", "database"):
            e = self._db_entries.get(key)
            if e:
                e.configure(state="disabled" if is_sqlite else "normal")
        self._sqlite_hint.configure(foreground="black" if is_sqlite else "gray")
        # Oracle Instant Client 行仅 Oracle 时显示
        if is_oracle:
            self._oracle_row.grid()
        else:
            self._oracle_row.grid_remove()
        if db_type == "mysql":
            self._db_vars["port"].set("3306")
        elif db_type == "oracle":
            self._db_vars["port"].set("1521")

    def _browse_sqlite_file(self):
        path = filedialog.asksaveasfilename(
            title=_("db.browse.sqlite.title"),
            filetypes=[("SQLite", "*.db *.sqlite"), ("All", "*.*")],
            defaultextension=".db",
        )
        if path:
            self._db_vars["host"].set(path)

    def _get_ui_db_cfg(self) -> dict:
        db_type = self.db_type_var.get()
        cfg = {"type": db_type, "name": self._db_vars["conn_name"].get().strip()}
        if db_type == "sqlite":
            cfg["path"] = self._db_vars["host"].get().strip()
        else:
            cfg["host"] = self._db_vars["host"].get().strip()
            cfg["port"] = self._db_vars["port"].get().strip()
            cfg["user"] = self._db_vars["user"].get().strip()
            cfg["password"] = self._db_vars["password"].get()
            if db_type == "mysql":
                cfg["database"] = self._db_vars["database"].get().strip()
            else:
                cfg["service"] = self._db_vars["database"].get().strip()
                cfg["lib_dir"] = self._db_vars["lib_dir"].get().strip()
        return cfg

    def _apply_cfg_to_ui(self, cfg: dict):
        self.db_type_var.set(cfg.get("type", "sqlite"))
        self._on_db_type_change()
        self._db_vars["conn_name"].set(cfg.get("name", ""))
        if cfg.get("type") == "sqlite":
            self._db_vars["host"].set(cfg.get("path", ""))
        else:
            self._db_vars["host"].set(cfg.get("host", ""))
            self._db_vars["port"].set(cfg.get("port", ""))
            self._db_vars["user"].set(cfg.get("user", ""))
            self._db_vars["password"].set(cfg.get("password", ""))
            self._db_vars["database"].set(cfg.get("database", cfg.get("service", "")))
            self._db_vars["lib_dir"].set(cfg.get("lib_dir", ""))

    def _save_db_config(self):
        cfg = self._get_ui_db_cfg()
        name = cfg.get("name", "")
        if not name:
            messagebox.showwarning(_("common.title.hint"), _("db.save.no_name"))
            return
        self.cfg_mgr.save(name, cfg)
        self._refresh_saved_combo()
        self.saved_name_var.set(name)
        messagebox.showinfo(_("db.save.done.title"), _("db.save.done", name=name))

    def _load_db_config(self):
        name = self.saved_name_var.get()
        cfg = self.cfg_mgr.get(name)
        if cfg:
            self._apply_cfg_to_ui(cfg)

    def _delete_db_config(self):
        name = self.saved_name_var.get()
        if not name:
            return
        if messagebox.askyesno(_("db.delete.title"), _("db.delete.confirm", name=name)):
            self.cfg_mgr.delete(name)
            self._refresh_saved_combo()

    def _refresh_saved_combo(self):
        names = self.cfg_mgr.names()
        self.saved_combo["values"] = names
        if names and self.saved_name_var.get() not in names:
            self.saved_name_var.set(names[-1])

    def _load_last_saved_config(self):
        names = self.cfg_mgr.names()
        if names:
            self.saved_name_var.set(names[-1])
            self._load_db_config()

    def _test_connection(self):
        cfg = self._get_ui_db_cfg()
        self.conn_status_var.set(_("db.test.connecting"))
        self.update_idletasks()
        try:
            db = DBConnection(cfg)
            db.connect()
            db.close()
            self.conn_status_var.set(_("db.test.ok"))
            messagebox.showinfo(_("db.test.ok.title"), _("db.test.ok.msg"))
        except Exception as e:
            self.conn_status_var.set(_("db.test.fail"))
            messagebox.showerror(_("db.test.fail.title"), str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # Tab 2 — CSV 文件配置
    # ══════════════════════════════════════════════════════════════════════════
    def _build_csv_tab(self):
        outer = ttk.Frame(self.nb)
        self.nb.add(outer, text=_("tab.csv"))

        sub = ttk.Notebook(outer)
        sub.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        f1 = ttk.Frame(sub)
        sub.add(f1, text=_("tab.csv_file"))
        self._build_csv_file_config(f1)

        f2 = ttk.Frame(sub)
        sub.add(f2, text=_("tab.csv_tools"))
        self._build_csv_char_tools(f2)


    def _build_csv_file_config(self, f):
        """Sub-tab: file selection + parse config + file info (original content)."""
        # 文件选择
        file_lf = ttk.LabelFrame(f, text=_("csv.file_select.label"))
        file_lf.pack(fill=tk.X, padx=12, pady=(12, 4))
        _ui = _load_ui_state()
        self.csv_path_var = tk.StringVar(value=_ui.get("last_csv_path", ""))
        ttk.Entry(file_lf, textvariable=self.csv_path_var, width=68).pack(side=tk.LEFT, padx=6, pady=8)
        ttk.Button(file_lf, text=_("common.btn.browse"), command=self._browse_csv).pack(side=tk.LEFT, padx=4)

        # 解析配置（含标题行、忽略行、编码；确认后统一生效）
        # 变量初始化（列分隔符/引用字符 Widget 在数据预览 tab 中创建）
        self.delimiter_var = tk.StringVar(value="|")
        self.quotechar_var = tk.StringVar(value='"')

        opt_lf = ttk.LabelFrame(f, text=_("csv.opt.label"))
        opt_lf.pack(fill=tk.X, padx=12, pady=4)

        # 编码行
        enc_row = ttk.Frame(opt_lf)
        enc_row.pack(fill=tk.X, padx=8, pady=(8, 2))
        ttk.Label(enc_row, text=_("csv.encoding.label")).pack(side=tk.LEFT)
        ttk.Entry(enc_row, textvariable=self.csv_encoding_var,
                  foreground="blue", width=20).pack(side=tk.LEFT, padx=4)

        # 标题行行
        hdr_row = ttk.Frame(opt_lf)
        hdr_row.pack(fill=tk.X, padx=8, pady=2)
        ttk.Checkbutton(hdr_row, text=_("csv.header.check"),
                        variable=self.has_header_var,
                        command=self._on_header_toggle).pack(side=tk.LEFT)
        ttk.Label(hdr_row, text=_("csv.header.row_label")).pack(side=tk.LEFT, padx=(16, 2))
        self.header_row_entry = ttk.Entry(hdr_row, textvariable=self.header_row_var, width=5)
        self.header_row_entry.pack(side=tk.LEFT)
        self.header_row_entry.configure(state=tk.DISABLED)

        # 忽略前/后行
        skip_row1 = ttk.Frame(opt_lf)
        skip_row1.pack(fill=tk.X, padx=8, pady=2)
        ttk.Label(skip_row1, text=_("csv.skip_head.label")).pack(side=tk.LEFT)
        self.skip_head_var = tk.StringVar(value="0")
        ttk.Entry(skip_row1, textvariable=self.skip_head_var, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(skip_row1, text=_("csv.skip_head.unit") + "   " + _("csv.skip_tail.label")).pack(side=tk.LEFT, padx=(4, 0))
        self.skip_tail_var = tk.StringVar(value="0")
        ttk.Entry(skip_row1, textvariable=self.skip_tail_var, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(skip_row1, text=_("csv.skip_tail.unit")).pack(side=tk.LEFT)

        # 忽略中间行
        skip_row2 = ttk.Frame(opt_lf)
        skip_row2.pack(fill=tk.X, padx=8, pady=2)
        ttk.Label(skip_row2, text=_("csv.skip_middle.label")).pack(side=tk.LEFT)
        self.skip_middle_var = tk.StringVar(value="")
        ttk.Entry(skip_row2, textvariable=self.skip_middle_var, width=40).pack(
            side=tk.LEFT, padx=4)
        ttk.Label(skip_row2, text=_("csv.skip_middle.hint"), foreground="gray").pack(side=tk.LEFT)

        # 确认行
        confirm_row = ttk.Frame(opt_lf)
        confirm_row.pack(fill=tk.X, padx=8, pady=(6, 8))
        ttk.Button(confirm_row, text=_("csv.confirm.btn"), command=self._confirm_skip).pack(side=tk.LEFT)
        self.skip_status_var = tk.StringVar(value=_("csv.confirm.pending"))
        ttk.Label(confirm_row, textvariable=self.skip_status_var,
                  foreground="blue").pack(side=tk.LEFT, padx=8)

        # 文件信息预览
        info_lf = ttk.LabelFrame(f, text=_("csv.info.label"))
        info_lf.pack(fill=tk.X, padx=12, pady=4)
        self.file_info_var = tk.StringVar(value=_("csv.info.no_file"))
        ttk.Label(info_lf, textvariable=self.file_info_var, foreground="gray").pack(
            anchor=tk.W, padx=8, pady=6)

        # 文件工具区
        tool_lf = ttk.LabelFrame(f, text=_("csv.tools.label"))
        tool_lf.pack(fill=tk.X, padx=12, pady=4)

        # 工具1：统计行数
        row_cnt = ttk.Frame(tool_lf)
        row_cnt.pack(fill=tk.X, padx=8, pady=(8, 2))
        ttk.Label(row_cnt, text=_("csv.tools.count.label")).pack(side=tk.LEFT)
        ttk.Button(row_cnt, text=_("csv.tools.count.btn"), command=self._ft_count_lines).pack(side=tk.LEFT, padx=6)
        self._ft_count_status = tk.StringVar(value="")
        ttk.Label(row_cnt, textvariable=self._ft_count_status, foreground="blue").pack(side=tk.LEFT, padx=4)

        ttk.Separator(tool_lf, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=8, pady=4)

        # 工具2：Peek 预览
        row_peek = ttk.Frame(tool_lf)
        row_peek.pack(fill=tk.X, padx=8, pady=2)
        ttk.Label(row_peek, text=_("csv.tools.peek.label")).pack(side=tk.LEFT)
        ttk.Label(row_peek, text=_("csv.tools.peek.n_label")).pack(side=tk.LEFT, padx=(8, 2))
        self._ft_peek_n_var = tk.StringVar(value="100")
        ttk.Entry(row_peek, textvariable=self._ft_peek_n_var, width=6).pack(side=tk.LEFT)
        ttk.Button(row_peek, text=_("csv.tools.peek.btn"), command=self._ft_peek).pack(side=tk.LEFT, padx=6)

        ttk.Separator(tool_lf, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=8, pady=4)

        # 工具3：文件拆分
        row_sp1 = ttk.Frame(tool_lf)
        row_sp1.pack(fill=tk.X, padx=8, pady=2)
        ttk.Label(row_sp1, text=_("csv.tools.split.label")).pack(side=tk.LEFT)
        self._ft_split_mode = tk.StringVar(value="lines")
        ttk.Radiobutton(row_sp1, text=_("csv.tools.split.lines"), variable=self._ft_split_mode,
                        value="lines").pack(side=tk.LEFT, padx=(8, 2))
        ttk.Radiobutton(row_sp1, text=_("csv.tools.split.size"), variable=self._ft_split_mode,
                        value="size").pack(side=tk.LEFT, padx=2)
        self._ft_split_val_var = tk.StringVar(value="1000000")
        ttk.Entry(row_sp1, textvariable=self._ft_split_val_var, width=10).pack(side=tk.LEFT, padx=4)
        ttk.Label(row_sp1, text=_("csv.tools.split.unit"), foreground="gray").pack(side=tk.LEFT)
        ttk.Button(row_sp1, text=_("csv.tools.split.btn"), command=self._ft_split).pack(side=tk.LEFT, padx=10)
        self._ft_split_status = tk.StringVar(value="")
        ttk.Label(row_sp1, textvariable=self._ft_split_status, foreground="blue").pack(side=tk.LEFT, padx=4)
        ttk.Frame(tool_lf).pack(pady=2)  # bottom padding

    def _browse_csv(self):
        path = filedialog.askopenfilename(
            title=_("csv.browse.title"),
            filetypes=[(_("csv.browse.filetype"), "*.csv *.txt"), ("All", "*.*")],
        )
        if path:
            self.csv_path_var.set(path)
            _save_ui_state({"last_csv_path": path})
            self._redetect_encoding()
            size = os.path.getsize(path)
            self.file_info_var.set(
                _("csv.info.path_size", path=path, size=size, mb=size / 1024 / 1024)
            )

    def _redetect_encoding(self):
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("csv.browse.no_file"))
            return
        enc = detect_encoding(path)
        self.csv_encoding_var.set(enc)

    def _on_header_toggle(self):
        """标题行勾选框切换时，启用/禁用行号输入框。"""
        state = tk.NORMAL if self.has_header_var.get() else tk.DISABLED
        self.header_row_entry.configure(state=state)

    def _confirm_skip(self):
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("csv.confirm.no_file"))
            return

        # 1. 验证标题行行号
        has_header = self.has_header_var.get()
        header_row = 0
        if has_header:
            hr_str = self.header_row_var.get().strip()
            if not hr_str or not hr_str.isdigit() or int(hr_str) < 1:
                messagebox.showwarning(_("common.title.hint"), _("csv.confirm.header_hint"))
                return
            header_row = int(hr_str)

        # 2. 解析忽略行数
        try:
            h = max(0, int(self.skip_head_var.get() or 0))
            t = max(0, int(self.skip_tail_var.get() or 0))
        except ValueError:
            messagebox.showwarning(_("common.title.hint"), _("csv.confirm.skip_int"))
            return
        middle_text = self.skip_middle_var.get().strip()
        try:
            middle_set = parse_ignore_ranges(middle_text) if middle_text else set()
        except Exception:
            messagebox.showwarning(_("common.title.hint"), _("csv.confirm.skip_fmt"))
            return

        # 3. 编码：优先使用用户手动填写的值，为空才自动检测
        enc = self.csv_encoding_var.get().strip() or detect_encoding(path)
        self.csv_encoding_var.set(enc)

        # 4. 应用设置
        self._skip_head = h
        self._skip_tail = t
        self._skip_middle = middle_set

        parts = []
        if h > 0:
            parts.append(_("csv.confirm.skip_head", n=h))
        if t > 0:
            parts.append(_("csv.confirm.skip_tail", n=t))
        if middle_set:
            parts.append(_("csv.confirm.skip_mid", n=len(middle_set)))
        skip_desc = "、".join(parts) if parts else _("csv.confirm.no_skip")
        self.skip_status_var.set(_("csv.confirm.done", enc=enc, skip_desc=skip_desc))

        # 5. 后台读取边界行（支持 20 GB 大文件）
        def _read_boundary():
            try:
                # 标题行内容
                header_content = None
                if has_header and header_row > 0:
                    raw = _fu_read_line_at(path, enc, header_row - 1)  # 0-based
                    if raw is not None:
                        header_content = raw[:200]

                # 第一条数据行：跳过 head、middle、标题行
                skip_set = set(middle_set)
                if has_header and header_row > 0:
                    skip_set.add(header_row)
                first_data = None
                first_data_lineno = None
                with open(path, encoding=enc, errors='replace') as fh:
                    for idx, line in enumerate(fh, start=1):
                        if idx <= h or idx in skip_set:
                            continue
                        first_data = line.rstrip('\r\n')[:200]
                        first_data_lineno = idx
                        break

                # 最后一条数据行：反向读取，跳过 tail、middle
                last_data = None
                total_approx = _fu_count_lines(path)
                CHUNK = 512 * 1024
                buf = b''
                pos = os.path.getsize(path)
                needed = t + 50 + len(middle_set)
                with open(path, 'rb') as fh:
                    while pos > 0 and buf.count(b'\n') < needed + 1:
                        read_size = min(CHUNK, pos)
                        pos -= read_size
                        fh.seek(pos)
                        buf = fh.read(read_size) + buf
                all_tail = buf.decode(enc, errors='replace').splitlines()
                while all_tail and all_tail[-1].strip() == '':
                    all_tail.pop()
                skipped = 0
                for rev_line in reversed(all_tail):
                    approx_lineno = total_approx - skipped
                    if t > 0 and approx_lineno > total_approx - t:
                        skipped += 1
                        continue
                    if approx_lineno in middle_set:
                        skipped += 1
                        continue
                    last_data = rev_line[:200]
                    break

                # ── 逐行原始字节编码检测，打印到控制台 ────────────────────────
                def _raw_line_at(fpath, idx_0based):
                    with open(fpath, 'rb') as fh:
                        for i, line in enumerate(fh):
                            if i == idx_0based:
                                return line.rstrip(b'\r\n')
                    return None

                enc_samples = []
                if has_header and header_row > 0:
                    raw = _raw_line_at(path, header_row - 1)
                    if raw:
                        enc_samples.append(("表头", header_row, raw))

                if first_data_lineno is not None:
                    raw = _raw_line_at(path, first_data_lineno - 1)
                    if raw:
                        enc_samples.append(("首行数据", first_data_lineno, raw))

                # 末行：从已读入的 buf 中提取原始字节（与 last_data 对应）
                buf_lines_b = buf.split(b'\n')
                while buf_lines_b and not buf_lines_b[-1].strip():
                    buf_lines_b.pop()
                skipped2 = 0
                for rev_raw in reversed(buf_lines_b):
                    rev_raw = rev_raw.rstrip(b'\r')
                    approx_ln = total_approx - skipped2
                    if t > 0 and approx_ln > total_approx - t:
                        skipped2 += 1
                        continue
                    if approx_ln in middle_set:
                        skipped2 += 1
                        continue
                    if rev_raw.strip():
                        enc_samples.append(("末行数据", approx_ln, rev_raw))
                    break

                print("─" * 60)
                print("[编码检测] 逐行采样：")
                for label, lineno, raw in enc_samples:
                    detected = detect_encoding_from_bytes(raw)
                    print(f"  {label}（第 {lineno} 行）: {detected:<12}  {raw[:80]}")
                print("─" * 60)
                # ─────────────────────────────────────────────────────────────

                def _show(hc=header_content, fd=first_data, fln=first_data_lineno, ld=last_data):
                    msg = _("csv.confirm.dialog.ok", enc=enc)
                    if hc is not None:
                        msg += _("csv.confirm.dialog.header", row=header_row, content=hc)
                    if fln is not None:
                        fd_display = fd if fd else _("csv.confirm.dialog.empty_line")
                        msg += _("csv.confirm.dialog.first_data", lineno=fln, content=fd_display)
                    else:
                        msg += _("csv.confirm.dialog.no_first")
                    last_content = ld if ld is not None else _("csv.confirm.dialog.no_data")
                    msg += _("csv.confirm.dialog.last_data", content=last_content)
                    messagebox.showinfo(_("csv.confirm.dialog.title"), msg)
                self.after(0, _show)
            except Exception:
                import traceback
                err = traceback.format_exc()
                self.after(0, lambda: messagebox.showerror(_("csv.confirm.read_fail"), err))

        threading.Thread(target=_read_boundary, daemon=True).start()

    def _get_skip_slice(self, rows: list) -> list:
        """Apply confirmed head/tail/middle skip. rows元素首项必须是 lineno（1-based）。"""
        h, t = self._skip_head, self._skip_tail
        end = len(rows) - t if t > 0 else len(rows)
        rows = rows[h:end]
        if self._skip_middle:
            rows = [r for r in rows if r[0] not in self._skip_middle]
        return rows

    # ── 文件工具后端 ──────────────────────────────────────────────────────────

    def _ft_get_path(self) -> str:
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("csv.confirm.no_file"))
            return ""
        return path

    # 工具1：统计行数（后台线程）
    def _ft_count_lines(self):
        path = self._ft_get_path()
        if not path:
            return
        self._ft_count_status.set(_("csv.count.counting"))
        def _run():
            try:
                count = _fu_count_lines(path)
                self.after(0, lambda c=count: (
                    self._ft_count_status.set(_("csv.count.result", n=c)),
                    messagebox.showinfo(_("csv.count.title"), _("csv.count.msg", path=path, n=c)),
                ))
            except Exception as e:
                self.after(0, lambda err=str(e): (
                    self._ft_count_status.set(_("csv.count.fail")),
                    messagebox.showerror(_("common.title.error"), err),
                ))
        threading.Thread(target=_run, daemon=True).start()

    # 工具2：Peek 预览（后台线程）
    def _ft_peek(self):
        path = self._ft_get_path()
        if not path:
            return
        try:
            n = max(1, int(self._ft_peek_n_var.get() or 100))
        except ValueError:
            messagebox.showwarning(_("common.title.hint"), _("csv.peek.invalid_n"))
            return

        def _run():
            try:
                enc = self.csv_encoding_var.get() or detect_encoding(path)
                file_size = os.path.getsize(path)
                head_lines = _fu_read_head(path, enc, n)
                tail_lines = _fu_read_tail(path, n, encoding=enc)

                out_path = path + '.peek.txt'
                with open(out_path, 'w', encoding='utf-8') as fh:
                    fh.write(_("csv.peek.file_info", path=path, size=file_size, enc=enc))
                    fh.write(f"{'─'*40}\n" + _("csv.peek.head_label", n=n) + f"\n{'─'*40}\n")
                    for i, ln in enumerate(head_lines, 1):
                        fh.write(f"{i:>6}: {ln}\n")
                    fh.write(f"\n{'─'*40}\n" + _("csv.peek.tail_label", n=n) + f"\n{'─'*40}\n")
                    for i, ln in enumerate(tail_lines, 1):
                        fh.write(f"{i:>6}: {ln}\n")

                def _show(hl=head_lines, tl=tail_lines, op=out_path):
                    top = tk.Toplevel(self)
                    top.title(_("csv.peek.title", n=n))
                    top.geometry("900x600")
                    st = scrolledtext.ScrolledText(top, font=("Consolas", 9), wrap=tk.NONE)
                    st.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
                    st.insert(tk.END, _("csv.peek.file_info", path=path, size=file_size, enc=enc))
                    st.insert(tk.END, f"\n{'─'*60}\n" + _("csv.peek.head_label", n=len(hl)) + f"\n{'─'*60}\n")
                    for i, ln in enumerate(hl, 1):
                        st.insert(tk.END, f"{i:>6}: {ln[:300]}\n")
                    st.insert(tk.END, f"\n{'─'*60}\n" + _("csv.peek.tail_label", n=len(tl)) + f"\n{'─'*60}\n")
                    for i, ln in enumerate(tl, 1):
                        st.insert(tk.END, f"{i:>6}: {ln[:300]}\n")
                    st.insert(tk.END, _("csv.peek.saved", path=op))
                    st.configure(state=tk.DISABLED)
                    hsb = ttk.Scrollbar(top, orient=tk.HORIZONTAL, command=st.xview)
                    hsb.pack(side=tk.BOTTOM, fill=tk.X)
                    st.configure(xscrollcommand=hsb.set)
                self.after(0, _show)
            except Exception as e:
                import traceback
                err = traceback.format_exc()
                self.after(0, lambda: messagebox.showerror(_("csv.peek.fail.title"), err))
        threading.Thread(target=_run, daemon=True).start()

    # 工具3：文件拆分（后台线程）
    def _ft_split(self):
        path = self._ft_get_path()
        if not path:
            return
        mode = self._ft_split_mode.get()
        try:
            val = float(self._ft_split_val_var.get())
            if val <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning(_("common.title.hint"), _("csv.split.invalid_val"))
            return

        name, ext = os.path.splitext(os.path.basename(path))
        ext = ext or '.txt'
        out_dir = os.path.dirname(os.path.abspath(path))
        file_size = os.path.getsize(path)
        self._ft_split_status.set(_("csv.split.running"))

        def _run():
            try:
                results = _fu_split_file(
                    path, out_dir, name, ext,
                    lines_per_file=int(val) if mode == 'lines' else 0,
                    max_bytes=int(val * 1024 * 1024) if mode == 'size' else 0,
                )

                def _show(res=results):
                    top = tk.Toplevel(self)
                    top.title(_("csv.split.done_title"))
                    top.geometry("800x400")
                    st = scrolledtext.ScrolledText(top, font=("Consolas", 9), wrap=tk.NONE)
                    st.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
                    st.insert(tk.END, _("csv.split.src_info", path=path, size=file_size))
                    st.insert(tk.END, _("csv.split.total", n=len(res)))
                    for fp, lc, bc in res:
                        st.insert(tk.END, _("csv.split.file_info", path=fp, lines=lc, mb=bc/1024/1024))
                    st.configure(state=tk.DISABLED)
                    hsb = ttk.Scrollbar(top, orient=tk.HORIZONTAL, command=st.xview)
                    hsb.pack(side=tk.BOTTOM, fill=tk.X)
                    st.configure(xscrollcommand=hsb.set)

                self.after(0, lambda: (
                    self._ft_split_status.set(_("csv.split.done", n=len(results))),
                    _show(),
                ))
            except FileExistsError as e:
                self.after(0, lambda err=str(e): (
                    self._ft_split_status.set(_("csv.split.conflict")),
                    messagebox.showerror(_("csv.split.conflict.title"), err),
                ))
            except Exception as e:
                import traceback
                err = traceback.format_exc()
                self.after(0, lambda: (
                    self._ft_split_status.set(_("csv.split.fail")),
                    messagebox.showerror(_("csv.split.fail.title"), err),
                ))
        threading.Thread(target=_run, daemon=True).start()

    # ── Sub-tab: 字符与分隔符探测 ─────────────────────────────────────────────

    def _build_csv_char_tools(self, f):
        """Sub-tab: Feature 1 (char search) + Feature 2 (auto detect delimiter)."""
        # ── Feature 1: 字符查找 ────────────────────────────────────────────────
        f1_lf = ttk.LabelFrame(f, text=_("csv.char.f1.label"))
        f1_lf.pack(fill=tk.X, padx=12, pady=(10, 4))

        row1 = ttk.Frame(f1_lf)
        row1.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(row1, text=_("csv.char.search.label")).pack(side=tk.LEFT)
        self.char_search_var = tk.StringVar()
        ttk.Entry(row1, textvariable=self.char_search_var, width=12).pack(side=tk.LEFT, padx=4)

        ttk.Label(row1, text=_("csv.char.scan.label")).pack(side=tk.LEFT)
        self.char_scan_full_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(row1, text=_("csv.char.full.label"), variable=self.char_scan_full_var).pack(side=tk.LEFT, padx=2)
        ttk.Label(row1, text=_("csv.char.sample_mb.label")).pack(side=tk.LEFT, padx=(8, 2))
        self.char_sample_mb_var = tk.StringVar(value="10")
        ttk.Entry(row1, textvariable=self.char_sample_mb_var, width=5).pack(side=tk.LEFT)

        self.char_search_btn = ttk.Button(row1, text=_("csv.char.btn"), command=self._start_char_search)
        self.char_search_btn.pack(side=tk.LEFT, padx=10)
        self.char_search_pb = ttk.Progressbar(row1, mode="determinate", length=160)
        self.char_search_pb.pack(side=tk.LEFT)

        self.char_search_result_var = tk.StringVar(value="")
        ttk.Label(f1_lf, textvariable=self.char_search_result_var, foreground="blue",
                  wraplength=800, justify=tk.LEFT).pack(anchor=tk.W, padx=8, pady=(0, 6))

        # ── Feature 2: 自动探测分隔符 ──────────────────────────────────────────
        f2_lf = ttk.LabelFrame(f, text=_("csv.detect.f2.label"))
        f2_lf.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)

        row2 = ttk.Frame(f2_lf)
        row2.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(row2, text=_("csv.detect.combo.label")).pack(side=tk.LEFT)
        self.combo_chars_var = tk.StringVar(value="!~^|=&")
        ttk.Entry(row2, textvariable=self.combo_chars_var, width=18).pack(side=tk.LEFT, padx=4)
        ttk.Label(row2, text=_("csv.detect.combo.hint"), foreground="gray").pack(side=tk.LEFT)

        ttk.Label(row2, text=_("csv.detect.mb.label")).pack(side=tk.LEFT, padx=(8, 2))
        self.auto_detect_mb_var = tk.StringVar(value="10")
        ttk.Entry(row2, textvariable=self.auto_detect_mb_var, width=5).pack(side=tk.LEFT)
        self.auto_detect_full_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(row2, text=_("csv.detect.full.label"), variable=self.auto_detect_full_var).pack(side=tk.LEFT, padx=4)

        self.auto_detect_btn = ttk.Button(row2, text=_("csv.detect.btn"), command=self._start_auto_detect)
        self.auto_detect_btn.pack(side=tk.LEFT, padx=10)

        row3 = ttk.Frame(f2_lf)
        row3.pack(fill=tk.X, padx=8, pady=(0, 4))
        self.auto_detect_pb = ttk.Progressbar(row3, mode="determinate", length=300)
        self.auto_detect_pb.pack(side=tk.LEFT)
        self.auto_detect_status_var = tk.StringVar(value="")
        ttk.Label(row3, textvariable=self.auto_detect_status_var, foreground="gray").pack(
            side=tk.LEFT, padx=8)

        self.auto_detect_result = scrolledtext.ScrolledText(
            f2_lf, height=10, state=tk.DISABLED, font=("Consolas", 9))
        self.auto_detect_result.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 6))

    # ── Feature 1 实现 ────────────────────────────────────────────────────────

    def _start_char_search(self):
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("csv.char.no_file"))
            return
        pattern = self.char_search_var.get()
        if not pattern:
            messagebox.showwarning(_("common.title.hint"), _("csv.char.no_pattern"))
            return

        file_size = os.path.getsize(path)
        if self.char_scan_full_var.get():
            scan_size = file_size
        else:
            try:
                mb = max(1, int(self.char_sample_mb_var.get() or 10))
            except ValueError:
                mb = 10
            scan_size = min(mb * 1024 * 1024, file_size)

        self.char_search_btn.configure(state="disabled")
        self.char_search_pb["value"] = 0
        self.char_search_result_var.set(_("csv.char.scanning"))

        threading.Thread(
            target=self._do_char_search,
            args=(path, pattern, scan_size, file_size),
            daemon=True,
        ).start()

    def _do_char_search(self, path, pattern_str, scan_size, file_size):
        try:
            try:
                pattern = pattern_str.encode('latin-1')
            except UnicodeEncodeError:
                pattern = pattern_str.encode('utf-8')

            overlap = max(0, len(pattern) - 1)
            READ_BUF = 8 * 1024 * 1024
            count = 0
            done = 0
            tail = b''

            with open(path, 'rb') as fh:
                while done < scan_size:
                    raw = fh.read(min(READ_BUF, scan_size - done))
                    if not raw:
                        break
                    buf = tail + raw
                    count += buf.count(pattern)
                    tail = raw[-overlap:] if overlap > 0 else b''
                    done += len(raw)
                    pct = done / scan_size * 100
                    self.after(0, lambda p=pct: self.char_search_pb.configure(value=p))

            is_sample = scan_size < file_size
            note = _("csv.char.sample_note", mb=scan_size // 1024 // 1024) if is_sample else ""
            if count == 0:
                msg = _("csv.char.result.safe", pattern=pattern_str, note=note)
            else:
                msg = _("csv.char.result.found", pattern=pattern_str, count=count, note=note)

            self.after(0, lambda m=msg: (
                self.char_search_pb.configure(value=100),
                self.char_search_result_var.set(m),
                self.char_search_btn.configure(state="normal"),
            ))
        except Exception as e:
            err = str(e)
            self.after(0, lambda: (
                self.char_search_result_var.set(_("csv.char.err", err=err)),
                self.char_search_btn.configure(state="normal"),
            ))

    # ── Feature 2 实现 ────────────────────────────────────────────────────────

    def _start_auto_detect(self):
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("csv.detect.no_file"))
            return

        combo_input = self.combo_chars_var.get().strip()
        combo_chars = list(dict.fromkeys(combo_input))  # 去重保序

        file_size = os.path.getsize(path)
        if self.auto_detect_full_var.get():
            scan_size = file_size
        else:
            try:
                mb = max(1, int(self.auto_detect_mb_var.get() or 10))
            except ValueError:
                mb = 10
            scan_size = min(mb * 1024 * 1024, file_size)

        self.auto_detect_btn.configure(state="disabled")
        self.auto_detect_pb["value"] = 0
        self.auto_detect_status_var.set(_("csv.detect.scanning"))
        self._auto_detect_result_clear()

        threading.Thread(
            target=self._do_auto_detect,
            args=(path, combo_chars, scan_size, file_size),
            daemon=True,
        ).start()

    def _do_auto_detect(self, path, combo_chars, scan_size, file_size):
        try:
            import itertools

            # 固定单字符候选集
            single_chars = list('!`=;&%+^@\\|/#~')
            single_patterns = [(c, c.encode('latin-1')) for c in single_chars]

            # 用户指定候选字符生成所有2字符组合
            combo_patterns = []
            for a, b in itertools.product(combo_chars, repeat=2):
                s = a + b
                try:
                    combo_patterns.append((s, s.encode('latin-1')))
                except UnicodeEncodeError:
                    pass

            all_patterns = single_patterns + combo_patterns
            if not all_patterns:
                self.after(0, lambda: (
                    self.auto_detect_status_var.set(_("csv.detect.no_pattern")),
                    self.auto_detect_btn.configure(state="normal"),
                ))
                return

            max_pat_len = max(len(p) for _, p in all_patterns)
            overlap = max_pat_len - 1
            counts = {p: 0 for _, p in all_patterns}

            READ_BUF = 8 * 1024 * 1024
            done = 0
            tail = b''

            with open(path, 'rb') as fh:
                while done < scan_size:
                    raw = fh.read(min(READ_BUF, scan_size - done))
                    if not raw:
                        break
                    buf = tail + raw
                    for _, pat in all_patterns:
                        counts[pat] += buf.count(pat)
                    tail = raw[-overlap:] if overlap > 0 else b''
                    done += len(raw)
                    pct = done / scan_size * 100
                    self.after(0, lambda p=pct: self.auto_detect_pb.configure(value=p))

            is_sample = scan_size < file_size
            note = _("csv.detect.sample_note", mb=scan_size // 1024 // 1024) if is_sample else _("csv.detect.full_note")

            lines = [_("csv.detect.result.title", note=note), "=" * 60 + "\n"]

            # single-char results
            safe_single = [s for s, p in single_patterns if counts[p] == 0]
            unsafe_single = sorted(
                [(s, counts[p]) for s, p in single_patterns if counts[p] > 0],
                key=lambda x: x[1])

            lines.append(_("csv.detect.single.safe.hdr"))
            if safe_single:
                lines.append("  " + "  ".join(safe_single) + "\n")
            else:
                lines.append(_("csv.detect.single.none"))

            lines.append(_("csv.detect.single.unsafe.hdr"))
            for s, cnt in unsafe_single:
                lines.append(f"  {repr(s):<8}  {cnt:>14,}" + _("csv.detect.single.count_unit"))

            # combo-char results
            if combo_patterns:
                safe_combo = [s for s, p in combo_patterns if counts[p] == 0]
                unsafe_combo = sorted(
                    [(s, counts[p]) for s, p in combo_patterns if counts[p] > 0],
                    key=lambda x: x[1])

                safe_same = [s for s in safe_combo if len(s) == 2 and s[0] == s[1]]
                safe_diff = [s for s in safe_combo if len(s) != 2 or s[0] != s[1]]

                lines.append(_("csv.detect.combo.safe.hdr", n=len(safe_combo)))
                if safe_same:
                    lines.append(_("csv.detect.combo.repeat") + "  ".join(safe_same) + "\n")
                if safe_diff:
                    for i in range(0, len(safe_diff), 15):
                        lines.append("  " + "  ".join(safe_diff[i:i + 15]) + "\n")
                if not safe_combo:
                    lines.append(_("csv.detect.single.none"))

                if safe_combo:
                    best = safe_same[0] if safe_same else safe_diff[0]
                    lines.append(_("csv.detect.combo.best", r=repr(best)))
                    lines.append(f"     Python: SEP = {repr(best)}\n")
                    lines.append(f"             fields = line.rstrip('\\n').split(SEP)\n")

                lines.append(_("csv.detect.combo.unsafe.hdr"))
                for s, cnt in unsafe_combo[:10]:
                    lines.append(f"  {repr(s):<8}  {cnt:>14,}" + _("csv.detect.single.count_unit"))

            result_text = "".join(lines)
            self.after(0, lambda t=result_text: (
                self.auto_detect_pb.configure(value=100),
                self.auto_detect_status_var.set(_("csv.detect.done")),
                self.auto_detect_btn.configure(state="normal"),
                self._auto_detect_result_set(t),
            ))
        except Exception as e:
            err = str(e)
            self.after(0, lambda: (
                self.auto_detect_status_var.set(_("csv.detect.err", err=err)),
                self.auto_detect_btn.configure(state="normal"),
            ))

    def _auto_detect_result_clear(self):
        self.auto_detect_result.configure(state=tk.NORMAL)
        self.auto_detect_result.delete("1.0", tk.END)
        self.auto_detect_result.configure(state=tk.DISABLED)

    def _auto_detect_result_set(self, text):
        self.auto_detect_result.configure(state=tk.NORMAL)
        self.auto_detect_result.delete("1.0", tk.END)
        self.auto_detect_result.insert(tk.END, text)
        self.auto_detect_result.configure(state=tk.DISABLED)

    # ══════════════════════════════════════════════════════════════════════════
    # Tab 3 — 数据校验
    # ══════════════════════════════════════════════════════════════════════════
    def _build_validate_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text=_("tab.validate"))

        # ── 控制区 ────────────────────────────────────────────────────────────
        ctrl = ttk.Frame(f)
        ctrl.pack(fill=tk.X, padx=12, pady=(4, 4))
        self.val_btn = ttk.Button(ctrl, text=_("val.btn.start"), command=self._start_validation)
        self.val_btn.pack(side=tk.LEFT, padx=4)
        self.add_invalid_btn = ttk.Button(ctrl, text=_("val.btn.add_ignore"),
                                          command=self._add_invalid_to_ignore, state="disabled")
        self.add_invalid_btn.pack(side=tk.LEFT, padx=8)
        self.val_status_var = tk.StringVar(value=_("val.status.waiting"))
        ttk.Label(ctrl, textvariable=self.val_status_var).pack(side=tk.LEFT, padx=10)

        self.val_pb = ttk.Progressbar(f, mode="determinate")
        self.val_pb.pack(fill=tk.X, padx=12, pady=4)

        stat = ttk.Frame(f)
        stat.pack(fill=tk.X, padx=12, pady=2)
        for label, attr, color in [
            (_("val.stat.std_cols"), "std_cols_var", "blue"),
            (_("val.stat.total"), "total_rows_var", "black"),
            (_("val.stat.data"), "data_rows_var", "black"),
            (_("val.stat.invalid"), "invalid_rows_var", "red"),
        ]:
            ttk.Label(stat, text=label).pack(side=tk.LEFT, padx=(12, 2))
            var = tk.StringVar(value="-")
            setattr(self, attr, var)
            ttk.Label(stat, textvariable=var, foreground=color, width=8).pack(side=tk.LEFT)

        ttk.Label(f, text=_("val.log.label")).pack(anchor=tk.W, padx=12, pady=(8, 2))
        self.val_log = scrolledtext.ScrolledText(f, height=12, state=tk.DISABLED,
                                                  font=("Consolas", 9))
        self.val_log.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 4))

        sum_lf = ttk.LabelFrame(f, text=_("val.summary.label"))
        sum_lf.pack(fill=tk.X, padx=12, pady=(4, 8))
        self.val_summary = scrolledtext.ScrolledText(sum_lf, height=14, state=tk.DISABLED,
                                                      font=("Consolas", 9), wrap=tk.WORD)
        self.val_summary.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

    def _on_val_mode_change(self):
        if self._val_mode_var.get() == "fixedwidth":
            self._csv_config_row.pack_forget()
            self._fw_config_row.pack(fill=tk.X, padx=8, pady=(2, 6))
        else:
            self._fw_config_row.pack_forget()
            self._csv_config_row.pack(fill=tk.X, padx=8, pady=(2, 6))

    def _auto_detect_fw_positions(self):
        """读取忽略规则之外的前100行，自动探测固定宽度分割位置，写入分割位置文本框。"""
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("preview.fw.no_file"))
            return
        encoding = self.csv_encoding_var.get() or detect_encoding(path)
        self.csv_encoding_var.set(encoding)
        self.fw_auto_detect_btn.configure(state="disabled")
        threading.Thread(
            target=self._do_auto_detect_fw,
            args=(path, encoding),
            daemon=True,
        ).start()

    def _do_auto_detect_fw(self, path, encoding):
        try:
            SAMPLE = 100
            threshold = 0.70
            pos_count: dict = {}
            sampled: list = []

            head = self._skip_head
            middle = self._skip_middle

            with open(path, 'r', encoding=encoding, errors='replace') as fh:
                for lineno, raw_line in enumerate(fh, start=1):
                    if lineno <= head:
                        continue
                    if middle and lineno in middle:
                        continue
                    s = raw_line.rstrip('\r\n')
                    if not s:
                        continue
                    sampled.append(s)
                    if len(sampled) >= SAMPLE:
                        break

            if not sampled:
                self.after(0, lambda: (
                    messagebox.showwarning(_("common.title.hint"), _("preview.fw.no_rows")),
                    self.fw_auto_detect_btn.configure(state="normal"),
                ))
                return

            for s in sampled:
                for i in range(1, len(s)):
                    if s[i - 1] == ' ' and s[i] != ' ':
                        pos_count[i] = pos_count.get(i, 0) + 1

            n = len(sampled)
            recommended = sorted(
                pos for pos, cnt in pos_count.items()
                if cnt / n >= threshold
            )
            all_pos = [0] + recommended
            positions_str = ", ".join(str(p) for p in all_pos)

            def _done(ps=positions_str, ns=n):
                self._fw_positions_var.set(ps)
                self.fw_auto_detect_btn.configure(state="normal")
                messagebox.showinfo(
                    _("preview.fw.done.title"),
                    _("preview.fw.done.msg", n=ns, positions=ps)
                )

            self.after(0, _done)

        except Exception as e:
            err = str(e)
            self.after(0, lambda: (
                messagebox.showerror(_("preview.fw.fail.title"), err),
                self.fw_auto_detect_btn.configure(state="normal"),
            ))

    def _start_validation(self):
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("val.no_file"))
            return
        encoding = self.csv_encoding_var.get() or detect_encoding(path)
        self.csv_encoding_var.set(encoding)

        mode = self._val_mode_var.get()
        fw_positions = None

        if mode == "fixedwidth":
            positions_str = self._fw_positions_var.get().strip()
            if not positions_str:
                messagebox.showwarning(_("common.title.hint"), _("val.fw.no_positions"))
                return
            try:
                fw_positions = sorted(set(
                    int(x.strip()) for x in positions_str.split(",") if x.strip()
                ))
                if not fw_positions:
                    raise ValueError
            except ValueError:
                messagebox.showwarning(_("common.title.hint"), _("val.fw.bad_positions"))
                return

        self._val_log_clear()
        self.val_status_var.set(_("val.status.start"))
        self.val_pb.configure(mode="indeterminate")
        self.val_pb.start(20)
        self.val_btn.configure(state="disabled")
        for attr in ("std_cols_var", "total_rows_var", "data_rows_var", "invalid_rows_var"):
            getattr(self, attr).set("-")
        self.validation_result = None

        has_header = self.has_header_var.get()
        header_row = 0
        if has_header:
            hr_str = self.header_row_var.get().strip()
            if not hr_str or not hr_str.isdigit() or int(hr_str) < 1:
                messagebox.showwarning(_("common.title.hint"), _("val.header_hint"))
                self.val_btn.configure(state="normal")
                return
            header_row = int(hr_str)

        if mode == "fixedwidth":
            threading.Thread(
                target=self._do_fw_validation,
                args=(path, encoding, fw_positions, has_header, header_row),
                daemon=True,
            ).start()
        else:
            threading.Thread(
                target=self._do_validation,
                args=(path, encoding, self.delimiter_var.get() or "|",
                      self.quotechar_var.get() or '"', has_header, header_row),
                daemon=True,
            ).start()

    def _do_validation(self, path, encoding, delimiter, quotechar, has_header, header_row):
        """
        单遍流式校验：文件只读一次，全程 O(异常行数) 内存。
        std_cols 从表头行或首批数据行确定，无需扫完全文件。
        """
        from collections import deque
        try:
            h = self._skip_head
            t = self._skip_tail
            m = self._skip_middle

            def _filtered_iter(src_iter):
                """生成器：对原始 (lineno, row) 流应用 head/middle/tail 忽略规则。"""
                buf = deque()
                for lineno, row in src_iter:
                    if lineno <= h:
                        continue
                    if m and lineno in m:
                        continue
                    if t > 0:
                        buf.append((lineno, row))
                        if len(buf) > t:
                            yield buf.popleft()
                    else:
                        yield lineno, row
                # buf 中剩余的是尾部行，按规则丢弃，不 yield

            # ── 单遍流式：边读边校验 ──────────────────────────────────────────
            total_lines  = 0
            data_count   = 0
            header_seen  = False
            columns      = None
            std_cols     = None   # 从表头或首行数据确定，之后固定不变
            invalid: list = []

            def _raw_with_progress():
                nonlocal total_lines
                for lineno, row in enumerate(
                        read_rows(path, encoding, delimiter, quotechar), start=1):
                    total_lines += 1
                    if total_lines % 50_000 == 0:
                        n = total_lines
                        self.after(0, lambda n=n:
                            self.val_status_var.set(_("val.status.progress", n=n)))
                    yield lineno, row

            for lineno, row in _filtered_iter(_raw_with_progress()):
                # ── 表头行 ──────────────────────────────────────────────────
                if has_header and lineno == header_row:
                    header_seen = True
                    columns  = [c.strip() or f"col_{i+1}" for i, c in enumerate(row)]
                    std_cols = len(columns)   # 表头列数即为标准列数，立即确定
                    continue

                # ── 数据行 ──────────────────────────────────────────────────
                data_count += 1

                # 无表头时，用第一条数据行确定标准列数
                if std_cols is None:
                    std_cols = len(row)
                    columns  = [f"col_{i+1}" for i in range(std_cols)]

                if len(row) != std_cols:
                    invalid.append((lineno, len(row), row))

            if data_count == 0 and not header_seen:
                self.after(0, lambda: (
                    messagebox.showwarning(_("common.title.hint"), _("val.empty_file")),
                    self.val_btn.configure(state="normal"),
                ))
                return

            if std_cols is None:
                std_cols = 0
            if not columns:
                columns = []

            header_count  = 1 if header_seen else 0
            ignored_count = total_lines - data_count - header_count
            total         = data_count + header_count

            result = {
                "std_cols":      std_cols,
                "columns":       columns,
                "invalid":       invalid,
                "total_lines":   total_lines,
                "total":         total,
                "data_count":    data_count,
                "ignored_count": ignored_count,
                "has_header":    has_header,
                "header_row":    header_row,
                "encoding":      encoding,
                "delimiter":     delimiter,
                "quotechar":     quotechar,
                # 供导入/导出重新流式读取
                "path":          path,
                "skip_head":     self._skip_head,
                "skip_tail":     self._skip_tail,
                "skip_middle":   frozenset(self._skip_middle),
            }
            self.after(0, lambda: self._on_validation_done(result, path))

        except Exception as e:
            err = str(e)
            self.after(0, lambda: (
                self.val_pb.stop(),
                self.val_pb.configure(mode="determinate"),
                messagebox.showerror(_("val.err.title"), err),
                self.val_status_var.set(_("val.status.fail")),
                self.val_btn.configure(state="normal"),
            ))

    def _stream_data_rows(self, vr: dict):
        """
        根据校验结果 vr，重新流式读取文件，逐行 yield (lineno, row)。
        只产出数据行（已应用 head/middle/tail 忽略规则，已排除表头行）。
        不存入内存，适合大文件导入/导出。
        """
        from collections import deque
        path       = vr["path"]
        encoding   = vr["encoding"]
        h          = vr.get("skip_head", 0)
        t          = vr.get("skip_tail", 0)
        m          = vr.get("skip_middle", frozenset())
        has_header = vr.get("has_header", False)
        header_row = vr.get("header_row", 0)
        fw_pos     = vr.get("fw_positions")
        delimiter  = vr.get("delimiter", "|")
        quotechar  = vr.get("quotechar", '"')

        def _raw():
            if fw_pos:
                def split_fw(s):
                    return [s[fw_pos[i]:fw_pos[i+1] if i+1 < len(fw_pos) else len(s)].strip()
                            for i in range(len(fw_pos))]
                with open(path, 'r', encoding=encoding, errors='replace', newline='') as fh:
                    for ln, line in enumerate(fh, start=1):
                        yield ln, split_fw(line.rstrip('\n\r'))
            else:
                for ln, row in enumerate(
                        read_rows(path, encoding, delimiter, quotechar), start=1):
                    yield ln, row

        buf = deque()
        for lineno, row in _raw():
            if lineno <= h:
                continue
            if m and lineno in m:
                continue
            if has_header and lineno == header_row:
                continue
            if t > 0:
                buf.append((lineno, row))
                if len(buf) > t:
                    yield buf.popleft()
            else:
                yield lineno, row
        # buf 中剩余为尾部行，丢弃

    def _get_preview_col_names(self, expected_count: int):
        """
        若数据预览已加载且列数与 expected_count 一致，
        返回用户在预览 tab 中编辑后的列名列表，否则返回 None。
        """
        if not self._preview_columns:
            return None
        names = self.preview_grid.get_col_names()
        return names if len(names) == expected_count else None

    def _on_validation_done(self, result: dict, path: str):
        # 用预览 tab 中用户编辑的列名覆盖文件解析出的列名
        preview_names = self._get_preview_col_names(result["std_cols"])
        if preview_names:
            result["columns"] = preview_names
        self.validation_result = result
        invalid = result["invalid"]
        std_cols = result["std_cols"]
        total = result["total"]
        data_count = result["data_count"]

        self.val_pb.stop()
        self.val_pb.configure(mode="determinate")
        self.val_pb["value"] = 100
        self.std_cols_var.set(str(std_cols))
        self.total_rows_var.set(str(total))
        self.data_rows_var.set(str(data_count))
        self.invalid_rows_var.set(str(len(invalid)))
        self.val_btn.configure(state="normal")

        filename = os.path.basename(path)
        self._val_log_append(_("val.log.file", path=path))
        self._val_log_append(_("val.log.summary", std_cols=std_cols, total=total, data_count=data_count, invalid_count=len(invalid)))
        self._val_log_append("─" * 80 + "\n")

        if invalid:
            self._val_log_append(f"{_('val.log.hdr_lineno'):>10}  {_('val.log.hdr_colcnt'):>6}  {_('val.log.hdr_preview')}\n")
            self._val_log_append("─" * 80 + "\n")
            log_lines = [_("val.log.file_full", path=path, std_cols=std_cols)]
            for lineno, col_cnt, row in invalid:
                preview = delimiter_join(row, result["delimiter"])[:200]
                line = f"{lineno:>10}  {col_cnt:>6}  {preview}\n"
                self._val_log_append(line)
                log_lines.append(line)
                self.logger.warning(f"invalid row {lineno}: col_count={col_cnt}")

            # print invalid line number summary
            linenos_str = ",".join(str(ln) for ln, *_ in invalid)
            self._val_log_append(_("val.log.linenos", linenos=linenos_str))

            # 保存校验日志
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_path = os.path.join(LOG_DIR, f"validate_{ts}.log")
            with open(log_path, "w", encoding="utf-8") as lf:
                lf.writelines(log_lines)

            self.add_invalid_btn.configure(state="normal")
            self.val_status_var.set(_("val.status.done_invalid", n=len(invalid)))
            messagebox.showwarning(
                _("val.result.title"),
                _("val.result.invalid.msg", filename=filename, n=len(invalid), std_cols=std_cols, log_path=log_path),
            )
        else:
            self.add_invalid_btn.configure(state="disabled")
            self.val_status_var.set(_("val.status.pass"))
            messagebox.showinfo(_("val.result.title"), _("val.result.pass.msg", filename=filename))

        self._fill_val_summary(result, path)

    def _val_log_clear(self):
        self.val_log.configure(state=tk.NORMAL)
        self.val_log.delete("1.0", tk.END)
        self.val_log.configure(state=tk.DISABLED)

    def _val_log_append(self, text: str):
        self.val_log.configure(state=tk.NORMAL)
        self.val_log.insert(tk.END, text)
        self.val_log.see(tk.END)
        self.val_log.configure(state=tk.DISABLED)

    def _add_invalid_to_ignore(self):
        if not self.validation_result:
            return
        invalid = self.validation_result.get("invalid", [])
        if not invalid:
            messagebox.showinfo(_("val.add_ignore.title"), _("val.add_ignore.none"))
            return
        new_set = {lineno for lineno, *_ in invalid}
        merged = self._skip_middle | new_set
        normalized = normalize_ranges(merged)
        self._skip_middle = merged
        self.skip_middle_var.set(normalized)
        # 用自定义对话框显示，避免内容过长撑爆 messagebox 超出屏幕
        dlg = tk.Toplevel(self)
        dlg.title(_("val.add_ignore.title"))
        dlg.resizable(True, True)
        dlg.grab_set()

        ttk.Label(dlg, text=_("val.add_ignore.done", n=len(new_set)),
                  font=("", 10, "bold")).pack(padx=16, pady=(14, 4), anchor=tk.W)
        ttk.Label(dlg, text=_("val.add_ignore.current")).pack(padx=16, anchor=tk.W)

        txt_frame = ttk.Frame(dlg)
        txt_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(4, 8))
        txt = tk.Text(txt_frame, wrap=tk.WORD, height=12, width=72,
                      relief="solid", bd=1)
        sb = ttk.Scrollbar(txt_frame, orient=tk.VERTICAL, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        txt.insert("1.0", normalized)
        txt.configure(state=tk.DISABLED)

        ttk.Button(dlg, text=_("val.add_ignore.btn_ok"), command=dlg.destroy, width=10).pack(pady=(0, 12))
        dlg.update_idletasks()
        # 居中显示，限制最大高度不超过屏幕
        max_h = int(self.winfo_screenheight() * 0.8)
        w, h = dlg.winfo_width(), min(dlg.winfo_height(), max_h)
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")
        self.wait_window(dlg)

    def _fill_val_summary(self, result: dict, path: str):
        file_size = os.path.getsize(path)
        size_mb = file_size / 1024 / 1024
        total_lines = result.get("total_lines", result.get("total", 0))
        ignored = result.get("ignored_count", 0)
        data_count = result["data_count"]
        has_header = result.get("has_header", False)
        columns = result.get("columns", [])
        invalid_count = len(result.get("invalid", []))
        pass_count = data_count - invalid_count
        header_count = 1 if has_header else 0

        lines = [
            _("val.summary.size", mb=size_mb),
            _("val.summary.total", n=total_lines),
            _("val.summary.ignored", n=ignored),
            _("val.summary.has_header", yn=_("val.summary.yes") if has_header else _("val.summary.no")),
        ]
        if has_header and columns:
            lines.append(_("val.summary.header_info", cols=", ".join(columns)))
        lines += [
            _("val.summary.pass_count", n=pass_count),
            _("val.summary.invalid_count", n=invalid_count),
        ]

        eq_parts = [_("val.summary.eq_ignored", n=ignored)]
        if has_header:
            eq_parts.append(_("val.summary.eq_header"))
        eq_parts += [_("val.summary.eq_pass", n=pass_count), _("val.summary.eq_invalid", n=invalid_count)]
        eq_sum = ignored + header_count + pass_count + invalid_count
        eq_str = " + ".join(eq_parts) + f" = {eq_sum}"
        lines.append(_("val.summary.relation", eq=eq_str))
        if eq_sum == total_lines:
            lines.append(_("val.summary.ok"))
        else:
            lines.append(_("val.summary.mismatch", sum=eq_sum, total=total_lines))

        summary_text = "\n".join(lines)
        self.val_summary.configure(state=tk.NORMAL)
        self.val_summary.delete("1.0", tk.END)
        self.val_summary.insert(tk.END, summary_text)
        self.val_summary.configure(state=tk.DISABLED)

    # ── Feature 4: 固定宽度校验 ───────────────────────────────────────────────

    def _do_fw_validation(self, path, encoding, positions, has_header, header_row):
        """固定宽度单遍流式校验，文件只读一次，不缓存全量行数据。"""
        from collections import deque
        try:
            def split_fw(line_str):
                fields = []
                for i, start in enumerate(positions):
                    end = positions[i + 1] if i + 1 < len(positions) else len(line_str)
                    fields.append(line_str[start:end].strip())
                return fields

            expected_cols = len(positions)
            last_pos      = positions[-1]
            h = self._skip_head
            t = self._skip_tail
            m = self._skip_middle

            def _fw_filtered_iter(fh):
                """对固定宽度行流应用 head/middle/tail 忽略规则。"""
                buf = deque()
                for lineno, raw_line in enumerate(fh, start=1):
                    line_str = raw_line.rstrip('\n\r')
                    if lineno <= h:
                        continue
                    if m and lineno in m:
                        continue
                    if t > 0:
                        buf.append((lineno, line_str))
                        if len(buf) > t:
                            yield buf.popleft()
                    else:
                        yield lineno, line_str

            # ── 单遍流式：边读边校验，不存全量数据 ───────────────────────────
            total_lines  = 0
            data_count   = 0
            header_seen  = False
            columns      = None
            invalid: list = []

            with open(path, 'r', encoding=encoding, errors='replace', newline='') as fh:
                def _counted(src):
                    nonlocal total_lines
                    for lineno, raw_line in enumerate(src, start=1):
                        total_lines += 1
                        if total_lines % 50_000 == 0:
                            n = total_lines
                            self.after(0, lambda n=n:
                                self.val_status_var.set(_("val.status.progress", n=n)))
                        yield lineno, raw_line

                for lineno, line_str in _fw_filtered_iter(_counted(fh)):
                    if has_header and lineno == header_row:
                        header_seen = True
                        columns = [c or f"col_{i+1}" for i, c in enumerate(split_fw(line_str))]
                        continue
                    data_count += 1
                    if len(line_str) < last_pos:
                        invalid.append((lineno, len(line_str), split_fw(line_str)))

            if data_count == 0 and not header_seen:
                self.after(0, lambda: (
                    messagebox.showwarning(_("common.title.hint"), _("val.empty_file")),
                    self.val_btn.configure(state="normal"),
                ))
                return

            if not columns:
                columns = [f"col_{i+1}" for i in range(expected_cols)]

            header_count  = 1 if header_seen else 0
            ignored_count = total_lines - data_count - header_count
            total         = data_count + header_count

            result = {
                "std_cols":      expected_cols,
                "columns":       columns,
                "invalid":       invalid,
                "total_lines":   total_lines,
                "total":         total,
                "data_count":    data_count,
                "ignored_count": ignored_count,
                "has_header":    has_header,
                "header_row":    header_row,
                "encoding":      encoding,
                "delimiter":     "|",
                "quotechar":     '"',
                "fw_positions":  positions,
                # 供导入/导出重新流式读取
                "path":          path,
                "skip_head":     self._skip_head,
                "skip_tail":     self._skip_tail,
                "skip_middle":   frozenset(self._skip_middle),
            }
            self.after(0, lambda: self._on_fw_validation_done(result, path))

        except Exception as e:
            err = str(e)
            self.after(0, lambda: (
                self.val_pb.stop(),
                self.val_pb.configure(mode="determinate"),
                messagebox.showerror(_("val.err.title"), err),
                self.val_status_var.set(_("val.status.fail")),
                self.val_btn.configure(state="normal"),
            ))

    def _on_fw_validation_done(self, result: dict, path: str):
        # 用预览 tab 中用户编辑的列名覆盖文件解析出的列名
        preview_names = self._get_preview_col_names(result["std_cols"])
        if preview_names:
            result["columns"] = preview_names
        self.validation_result = result
        invalid = result["invalid"]  # list of (lineno, raw_len, fields)
        std_cols = result["std_cols"]
        total = result["total"]
        data_count = result["data_count"]
        positions = result["fw_positions"]

        self.val_pb.stop()
        self.val_pb.configure(mode="determinate")
        self.val_pb["value"] = 100
        self.std_cols_var.set(str(std_cols))
        self.total_rows_var.set(str(total))
        self.data_rows_var.set(str(data_count))
        self.invalid_rows_var.set(str(len(invalid)))
        self.val_btn.configure(state="normal")

        filename = os.path.basename(path)
        pos_str = ", ".join(str(p) for p in positions)
        self._val_log_append(_("val.log.file", path=path))
        self._val_log_append(_("val.log.fw.mode", pos_str=pos_str))
        self._val_log_append(_("val.log.fw.summary", std_cols=std_cols, total=total, data_count=data_count, invalid_count=len(invalid)))
        self._val_log_append("─" * 80 + "\n")

        if invalid:
            self._val_log_append(f"{_('val.log.fw.hdr_lineno'):>10}  {_('val.log.fw.hdr_rawlen'):>8}  {_('val.log.hdr_preview')}\n")
            self._val_log_append("─" * 80 + "\n")
            log_lines = [_("val.log.fw.file_full", path=path, pos_str=pos_str)]
            for lineno, raw_len, fields in invalid:
                preview = "|".join(str(v) for v in fields)[:200]
                line_str = f"{lineno:>10}  {raw_len:>8}  {preview}\n"
                self._val_log_append(line_str)
                log_lines.append(line_str)
                self.logger.warning(f"row {lineno}: length={raw_len}, required {positions[-1]}")

            linenos_str = ",".join(str(ln) for ln, *_ in invalid)
            self._val_log_append(_("val.log.fw.linenos", linenos=linenos_str))

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_path = os.path.join(LOG_DIR, f"validate_fw_{ts}.log")
            with open(log_path, "w", encoding="utf-8") as lf:
                lf.writelines(log_lines)

            self.add_invalid_btn.configure(state="normal")
            self.val_status_var.set(_("val.fw.status.done", n=len(invalid)))
            messagebox.showwarning(
                _("val.result.title"),
                _("val.fw.result.invalid.msg", filename=filename, pos_str=pos_str, n=len(invalid), last_pos=positions[-1], log_path=log_path),
            )
        else:
            self.add_invalid_btn.configure(state="disabled")
            self.val_status_var.set(_("val.fw.status.pass"))
            messagebox.showinfo(_("val.result.title"),
                _("val.fw.result.pass.msg", filename=filename, pos_str=pos_str))

        self._fill_val_summary(result, path)

    # ══════════════════════════════════════════════════════════════════════════
    # Tab 3 — 数据预览
    # ══════════════════════════════════════════════════════════════════════════
    def _build_preview_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text=_("tab.preview"))

        # ── 分割模式选择 ──────────────────────────────────────────────────────
        mode_lf = ttk.LabelFrame(f, text=_("preview.mode.label"))
        mode_lf.pack(fill=tk.X, padx=12, pady=(10, 4))

        mode_row = ttk.Frame(mode_lf)
        mode_row.pack(fill=tk.X, padx=8, pady=(6, 2))
        ttk.Radiobutton(mode_row, text=_("preview.mode.delimiter"), variable=self._val_mode_var,
                        value="delimiter", command=self._on_val_mode_change).pack(
            side=tk.LEFT, padx=4)
        ttk.Radiobutton(mode_row, text=_("preview.mode.fixedwidth"), variable=self._val_mode_var,
                        value="fixedwidth", command=self._on_val_mode_change).pack(
            side=tk.LEFT, padx=12)

        # 分隔符模式配置行（默认显示）
        self._csv_config_row = ttk.Frame(mode_lf)
        self._csv_config_row.pack(fill=tk.X, padx=8, pady=(2, 6))
        ttk.Label(self._csv_config_row, text=_("preview.delimiter.label")).pack(side=tk.LEFT)
        ttk.Entry(self._csv_config_row, textvariable=self.delimiter_var, width=8).pack(
            side=tk.LEFT, padx=4)
        ttk.Label(self._csv_config_row, text=_("preview.quotechar.label")).pack(side=tk.LEFT, padx=(12, 2))
        ttk.Entry(self._csv_config_row, textvariable=self.quotechar_var, width=5).pack(
            side=tk.LEFT)

        # 固定宽度模式配置行（默认隐藏）
        self._fw_config_row = ttk.Frame(mode_lf)
        self._fw_config_row.pack(fill=tk.X, padx=8, pady=(2, 6))
        ttk.Label(self._fw_config_row, text=_("preview.fw.pos.label")).pack(side=tk.LEFT)
        ttk.Entry(self._fw_config_row, textvariable=self._fw_positions_var, width=46).pack(
            side=tk.LEFT, padx=4)
        ttk.Label(self._fw_config_row, text=_("preview.fw.pos.hint"), foreground="gray").pack(
            side=tk.LEFT)
        self.fw_auto_detect_btn = ttk.Button(
            self._fw_config_row, text=_("preview.fw.auto.btn"), command=self._auto_detect_fw_positions)
        self.fw_auto_detect_btn.pack(side=tk.LEFT, padx=8)
        self._fw_config_row.pack_forget()  # 默认隐藏

        # 工具栏
        ctrl = ttk.Frame(f)
        ctrl.pack(fill=tk.X, padx=12, pady=(10, 4))
        ttk.Button(ctrl, text=_("preview.btn.load"), command=self._load_preview).pack(side=tk.LEFT, padx=4)
        ttk.Separator(ctrl, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        # 分页控件
        self.prev_page_btn = ttk.Button(ctrl, text=_("preview.page.prev"), command=self._preview_prev, state="disabled")
        self.prev_page_btn.pack(side=tk.LEFT, padx=2)
        self.preview_page_var = tk.StringVar(value="")
        ttk.Label(ctrl, textvariable=self.preview_page_var, width=14, anchor=tk.CENTER).pack(side=tk.LEFT, padx=4)
        self.next_page_btn = ttk.Button(ctrl, text=_("preview.page.next"), command=self._preview_next, state="disabled")
        self.next_page_btn.pack(side=tk.LEFT, padx=2)

        ttk.Separator(ctrl, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        self.preview_info_var = tk.StringVar(value="")
        ttk.Label(ctrl, textvariable=self.preview_info_var, foreground="gray").pack(side=tk.LEFT, padx=4)

        # _PreviewGrid（Canvas-based，奇偶列变色 + 可编辑表头）
        self.preview_grid = _PreviewGrid(f)
        self.preview_grid.pack(fill=tk.BOTH, expand=True, padx=12, pady=(4, 8))

        # 内部状态
        self._preview_all_rows: list = []       # 全部数据行（小文件模式）
        self._preview_columns: list = []        # 列名
        self._preview_page: int = 0             # 当前页（0-based）
        self._preview_page_size: int = 200      # 每页行数
        self._preview_lazy: bool = False        # 大文件流式模式
        self._preview_lazy_chunks: list = []    # 已加载的分块缓存
        self._preview_lazy_gen = None           # 行生成器（大文件模式）
        self._preview_lazy_done: bool = False   # 生成器已耗尽

    def _load_preview(self):
        path = self.csv_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning(_("common.title.hint"), _("preview.no_file"))
            return

        encoding = self.csv_encoding_var.get() or detect_encoding(path)
        self.csv_encoding_var.set(encoding)
        delimiter = self.delimiter_var.get() or "|"
        quotechar = self.quotechar_var.get() or '"'
        has_header = self.has_header_var.get()
        header_row = 0
        if has_header:
            hr_str = self.header_row_var.get().strip()
            if hr_str.isdigit() and int(hr_str) >= 1:
                header_row = int(hr_str)

        self.preview_info_var.set(_("preview.loading"))
        self.update_idletasks()

        threading.Thread(
            target=self._do_load_preview,
            args=(path, encoding, delimiter, quotechar, has_header, header_row),
            daemon=True,
        ).start()

    def _do_load_preview(self, path, encoding, delimiter, quotechar, has_header, header_row=0):
        try:
            file_size = os.path.getsize(path)
            lazy = file_size > 100 * 1024 * 1024   # > 100 MB

            h      = self._skip_head
            middle = self._skip_middle
            mode   = self._val_mode_var.get()

            # ── 固定宽度模式：行读取器 ─────────────────────────────────────────
            fw_positions = None
            if mode == "fixedwidth":
                pos_str = self._fw_positions_var.get().strip()
                if pos_str:
                    try:
                        fw_positions = sorted(set(
                            int(x.strip()) for x in pos_str.split(",") if x.strip()))
                    except ValueError:
                        fw_positions = None

            def _iter_rows_fw(positions):
                """生成器：逐行读取并按固定宽度分割，返回字段列表。"""
                with open(path, 'r', encoding=encoding, errors='replace', newline='') as fh:
                    for line in fh:
                        line_str = line.rstrip('\n\r')
                        fields = []
                        for i, start in enumerate(positions):
                            end = positions[i + 1] if i + 1 < len(positions) else len(line_str)
                            fields.append(line_str[start:end].strip())
                        yield fields

            def _iter_rows():
                if fw_positions:
                    yield from _iter_rows_fw(fw_positions)
                else:
                    yield from read_rows(path, encoding, delimiter, quotechar)

            if lazy:
                # ── 大文件流式模式：不读全文件，逐块加载 ──────────────────────
                # 先确定列名（读到标题行即止）
                columns = None
                if has_header and header_row > 0:
                    for lineno, row in enumerate(_iter_rows(), start=1):
                        if lineno == header_row:
                            columns = [c.strip() or f"col_{i+1}"
                                       for i, c in enumerate(row)]
                            break

                # 生成器：跳过 head / middle / header 行，不处理 tail
                def _make_gen():
                    for ln, row in enumerate(_iter_rows(), start=1):
                        if ln <= h:
                            continue
                        if middle and ln in middle:
                            continue
                        if has_header and header_row > 0 and ln == header_row:
                            continue
                        yield row

                gen = _make_gen()

                # 取第一块
                first_chunk = []
                for row in gen:
                    first_chunk.append(row)
                    if len(first_chunk) >= self._preview_page_size:
                        break

                if not first_chunk:
                    self.after(0, lambda: self.preview_info_var.set(_("preview.empty")))
                    return

                if columns is None:
                    if fw_positions:
                        columns = [f"col_{i+1}" for i in range(len(fw_positions))]
                    else:
                        max_cols = max(len(r) for r in first_chunk)
                        columns = [f"col_{i+1}" for i in range(max_cols)]

                self._preview_lazy = True
                self._preview_lazy_chunks = [first_chunk]
                self._preview_lazy_gen = gen
                self._preview_lazy_done = len(first_chunk) < self._preview_page_size
                self._preview_columns = columns
                self._preview_page = 0
                self.after(0, self._render_preview_page)

            else:
                # ── 小文件模式：全量加载 ───────────────────────────────────────
                rows_raw = [(lineno, row) for lineno, row in
                            enumerate(_iter_rows(), start=1)]
                rows_filtered = self._get_skip_slice(rows_raw)
                if not rows_filtered:
                    self.after(0, lambda: self.preview_info_var.set(_("preview.empty_small")))
                    return

                if has_header and header_row > 0:
                    header_match = [r for ln, r in rows_filtered if ln == header_row]
                    data_rows = [r for ln, r in rows_filtered if ln != header_row]
                    if header_match:
                        columns = [c.strip() or f"col_{i+1}"
                                   for i, c in enumerate(header_match[0])]
                    elif fw_positions:
                        columns = [f"col_{i+1}" for i in range(len(fw_positions))]
                    else:
                        columns = [f"col_{i+1}" for i in range(
                            max((len(r) for _, r in rows_filtered), default=0))]
                else:
                    data_rows = [r for _, r in rows_filtered]
                    if fw_positions:
                        columns = [f"col_{i+1}" for i in range(len(fw_positions))]
                    else:
                        max_cols = max((len(r) for r in data_rows), default=0)
                        columns = [f"col_{i+1}" for i in range(max_cols)]

                self._preview_lazy = False
                self._preview_all_rows = data_rows
                self._preview_columns = columns
                self._preview_page = 0
                self.after(0, self._render_preview_page)

        except Exception as e:
            self.after(0, lambda err=str(e): messagebox.showerror(_("preview.err.title"), err))
            self.after(0, lambda: self.preview_info_var.set(_("preview.load_fail")))

    def _fetch_next_lazy_chunk(self):
        """后台线程：从生成器取下一块数据并缓存。"""
        try:
            chunk = []
            for row in self._preview_lazy_gen:
                chunk.append(row)
                if len(chunk) >= self._preview_page_size:
                    break
            if chunk:
                self._preview_lazy_chunks.append(chunk)
                self._preview_page = len(self._preview_lazy_chunks) - 1
                if len(chunk) < self._preview_page_size:
                    self._preview_lazy_done = True
            else:
                self._preview_lazy_done = True
            self.after(0, self._render_preview_page)
        except Exception as e:
            self.after(0, lambda err=str(e): messagebox.showerror(_("preview.load_err.title"), err))
            self.after(0, lambda: self.preview_info_var.set(_("preview.load_fail")))

    def _render_preview_page(self):
        columns = self._preview_columns
        page = self._preview_page

        if self._preview_lazy:
            if page >= len(self._preview_lazy_chunks):
                return
            page_rows = self._preview_lazy_chunks[page]
            total_cached = sum(len(c) for c in self._preview_lazy_chunks)
            can_next = (page < len(self._preview_lazy_chunks) - 1) or not self._preview_lazy_done
            page_label = _("preview.page.streaming", page=page + 1)
            info_label = _("preview.info.rows_cached", page_rows=len(page_rows), total_cached=total_cached)
        else:
            all_rows = self._preview_all_rows
            ps = self._preview_page_size
            total = len(all_rows)
            total_pages = max(1, (total + ps - 1) // ps)
            start = page * ps
            end = min(start + ps, total)
            page_rows = all_rows[start:end]
            can_next = page < total_pages - 1
            page_label = _("preview.page.label_total", page=page + 1, total=total_pages)
            info_label = _("preview.info.rows_total", total=total, start=start + 1, end=end)

        # 取上次保存的列宽（切换页时保持宽度）
        saved_widths = self.preview_grid.get_col_widths() if self._preview_columns else {}
        self.preview_grid.set_data(columns, page_rows, saved_widths)

        self.preview_page_var.set(page_label)
        self.preview_info_var.set(info_label)
        self.prev_page_btn.configure(state="normal" if page > 0 else "disabled")
        self.next_page_btn.configure(state="normal" if can_next else "disabled")

    def _preview_prev(self):
        if self._preview_page > 0:
            self._preview_page -= 1
            self._render_preview_page()

    def _preview_next(self):
        if self._preview_lazy:
            if self._preview_page < len(self._preview_lazy_chunks) - 1:
                # 已缓存的页，直接跳转
                self._preview_page += 1
                self._render_preview_page()
            elif not self._preview_lazy_done:
                # 需要从文件读取下一块
                self.next_page_btn.configure(state="disabled")
                self.preview_info_var.set(_("preview.load_more"))
                threading.Thread(target=self._fetch_next_lazy_chunk, daemon=True).start()
        else:
            total = len(self._preview_all_rows)
            total_pages = max(1, (total + self._preview_page_size - 1) // self._preview_page_size)
            if self._preview_page < total_pages - 1:
                self._preview_page += 1
                self._render_preview_page()

    # ══════════════════════════════════════════════════════════════════════════
    # Tab 4 — 数据导入
    # ══════════════════════════════════════════════════════════════════════════
    def _build_import_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text=_("tab.import"))

        opt_lf = ttk.LabelFrame(f, text=_("import.opt.label"))
        opt_lf.pack(fill=tk.X, padx=12, pady=(12, 4))

        ttk.Label(opt_lf, text=_("import.table.label")).grid(row=0, column=0, sticky=tk.W, padx=8, pady=6)
        self.table_name_var = tk.StringVar()
        ttk.Entry(opt_lf, textvariable=self.table_name_var, width=32).grid(
            row=0, column=1, sticky=tk.W, padx=6, pady=6)

        ttk.Label(opt_lf, text=_("import.batch.label")).grid(row=0, column=2, sticky=tk.W, padx=(20, 6))
        self.batch_size_var = tk.StringVar(value="1000")
        ttk.Entry(opt_lf, textvariable=self.batch_size_var, width=10).grid(row=0, column=3, sticky=tk.W)

        self.create_table_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt_lf, text=_("import.create_table"),
                         variable=self.create_table_var).grid(
            row=1, column=0, columnspan=4, sticky=tk.W, padx=8, pady=3)

        self.truncate_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_lf, text=_("import.truncate"),
                         variable=self.truncate_var).grid(
            row=2, column=0, columnspan=4, sticky=tk.W, padx=8, pady=3)

        ttk.Label(opt_lf, text=_("import.error_mode.label")).grid(row=3, column=0, sticky=tk.W, padx=8, pady=6)
        self.error_mode_var = tk.StringVar(value="skip")
        ttk.Radiobutton(opt_lf, text=_("import.error_mode.skip"),
                         variable=self.error_mode_var, value="skip").grid(
            row=3, column=1, columnspan=3, sticky=tk.W)
        ttk.Radiobutton(opt_lf, text=_("import.error_mode.rollback"),
                         variable=self.error_mode_var, value="rollback").grid(
            row=4, column=1, columnspan=3, sticky=tk.W, padx=6)

        # 操作按钮
        btn_row = ttk.Frame(f)
        btn_row.pack(fill=tk.X, padx=12, pady=6)
        self.import_btn = ttk.Button(btn_row, text=_("import.btn.start"), command=self._start_import)
        self.import_btn.pack(side=tk.LEFT, padx=4)
        self.stop_btn = ttk.Button(btn_row, text=_("import.btn.stop"), command=self._stop_import_clicked,
                                   state="disabled")
        self.stop_btn.pack(side=tk.LEFT, padx=4)
        self.import_status_var = tk.StringVar(value="")
        ttk.Label(btn_row, textvariable=self.import_status_var).pack(side=tk.LEFT, padx=10)

        # 进度条
        self.imp_pb = ttk.Progressbar(f, mode="determinate")
        self.imp_pb.pack(fill=tk.X, padx=12, pady=4)

        # 计数器
        cnt = ttk.Frame(f)
        cnt.pack(fill=tk.X, padx=12)
        for label, attr, color in [
            (_("import.cnt.imported"), "imported_var", "green"),
            (_("import.cnt.skipped"), "skipped_var", "orange"),
            (_("import.cnt.errors"), "errors_var", "red"),
        ]:
            ttk.Label(cnt, text=label).pack(side=tk.LEFT, padx=(16, 2))
            var = tk.StringVar(value="0")
            setattr(self, attr, var)
            ttk.Label(cnt, textvariable=var, foreground=color, width=10).pack(side=tk.LEFT)

        ttk.Label(f, text=_("import.log.label")).pack(anchor=tk.W, padx=12, pady=(8, 2))
        self.imp_log = scrolledtext.ScrolledText(f, height=12, state=tk.DISABLED,
                                                  font=("Consolas", 9))
        self.imp_log.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 8))

    def _start_import(self):
        table = self.table_name_var.get().strip()
        if not table:
            messagebox.showwarning(_("common.title.hint"), _("import.no_table"))
            return
        if not self.validation_result:
            messagebox.showwarning(_("common.title.hint"), _("import.no_validate"))
            return

        db_cfg = self._get_ui_db_cfg()

        self._imp_log_clear()
        self.import_status_var.set(_("import.status.running"))
        self.imp_pb["value"] = 0
        self.imported_var.set("0")
        self.skipped_var.set("0")
        self.errors_var.set("0")
        self.ignored_error_types = set()
        self._stop_import.clear()
        self.import_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")

        threading.Thread(
            target=self._do_import,
            args=(db_cfg, table),
            daemon=True,
        ).start()

    def _do_import(self, db_cfg: dict, table: str):
        vr = self.validation_result
        columns: list = vr["columns"]
        std_cols: int = vr["std_cols"]
        error_mode: str = self.error_mode_var.get()
        create_table: bool = self.create_table_var.get()
        do_truncate: bool = self.truncate_var.get()
        try:
            batch_size = max(1, int(self.batch_size_var.get() or 1000))
        except ValueError:
            batch_size = 1000

        # ── 建立连接 ──────────────────────────────────────────────────────────
        try:
            db = DBConnection(db_cfg)
            db.connect()
        except Exception as e:
            self._finish_import_error(_("import.err.connect", e=e))
            return

        self._imp_log(_("import.log.connected", table=table))

        final_columns = list(columns)  # may be adjusted for column alignment

        try:
            # ── 表处理 ────────────────────────────────────────────────────────
            if db.table_exists(table):
                if do_truncate:
                    self._imp_log(_("import.log.truncate", table=table))
                    db.truncate_table(table)

                existing_cols = db.get_table_columns(table)
                csv_set = {c.lower() for c in columns}
                tbl_set = {c.lower() for c in existing_cols}
                if csv_set != tbl_set:
                    # 列不匹配，问用户是否对齐插入
                    answer = self._ask_from_thread(
                        _("import.col_mismatch.title"),
                        _("import.col_mismatch.msg", table=table, csv_cols=columns, tbl_cols=existing_cols),
                    )
                    if not answer:
                        db.close()
                        self.after(0, lambda: (
                            self.import_status_var.set(_("import.status.cancelled")),
                            self.import_btn.configure(state="normal"),
                        ))
                        return
                    # 对齐：只保留两边都有的列
                    tbl_lower = {c.lower(): c for c in existing_cols}
                    final_columns = [c for c in columns if c.lower() in tbl_lower]
                    if not final_columns:
                        db.close()
                        self._finish_import_error(
                            _("import.err.no_common_cols", table=table, csv_cols=columns, tbl_cols=existing_cols)
                        )
                        return
                    self._imp_log(_("import.log.aligned", cols=final_columns))
            else:
                if create_table:
                    self._imp_log(_("import.log.create", table=table, cols=columns))
                    db.create_table(table, columns)
                else:
                    db.close()
                    self._finish_import_error(_("import.err.no_table", table=table))
                    return

            # ── 构造 SQL ──────────────────────────────────────────────────────
            col_indices = [columns.index(c) for c in final_columns]
            col_sql = ", ".join([db.quote(c) for c in final_columns])
            ph_sql = db.placeholders(final_columns)
            insert_sql = f"INSERT INTO {db.quote(table)} ({col_sql}) VALUES ({ph_sql})"
            self._imp_log(f"INSERT SQL: {insert_sql}\n")
            self._imp_log("─" * 60 + "\n")

            data_rows  = self._stream_data_rows(vr)   # 流式读取，不占内存
            total_data = vr["data_count"]
            imported = skipped = errors = 0
            batch: list = []

            def log_batch_sql(first_row, batch_end):
                if isinstance(first_row, dict):
                    vals_str = ", ".join(f"{k}={repr(v)}" for k, v in first_row.items())
                else:
                    vals_str = ", ".join(repr(v) for v in first_row)
                msg = _("import.log.batch_hdr", batch=batch_end, total=total_data, sql=insert_sql, vals=vals_str)
                print(msg, flush=True)
                self._imp_log(msg)

            def flush(b):
                nonlocal imported
                log_batch_sql(b[0], imported + len(b))
                cur = db.cursor()
                if db_cfg.get("type") == "oracle":
                    for rv in b:
                        cur.execute(insert_sql, rv)
                else:
                    cur.executemany(insert_sql, b)
                db.commit()
                imported += len(b)

            for idx, (lineno, row) in enumerate(data_rows):
                # 列数不符处理
                if len(row) != std_cols:
                    if error_mode == "rollback":
                        db.rollback()
                        db.close()
                        self._finish_import_error(
                            _("import.err.col_mismatch_row", lineno=lineno, std_cols=std_cols, col_cnt=len(row))
                        )
                        return
                    else:
                        skipped += 1
                        self.logger.warning(f"跳过行 {lineno}: 列数={len(row)}")
                        continue

                values = [row[i] if i < len(row) else "" for i in col_indices]
                row_val = db.make_row(values, final_columns)

                if error_mode == "rollback":
                    batch.append(row_val)
                    if len(batch) >= batch_size:
                        try:
                            flush(batch)
                            batch = []
                        except Exception as e:
                            db.rollback()
                            db.close()
                            self._finish_import_error(_("import.err.write_fail", lineno=lineno, e=e))
                            return
                else:  # skip
                    try:
                        db.cursor().execute(insert_sql, row_val)
                        imported += 1
                        if imported % batch_size == 0:
                            log_batch_sql(row_val, imported)
                            db.commit()
                    except Exception as e:
                        errors += 1
                        etype = type(e).__name__
                        self.logger.error(f"行 {lineno} 写入失败 [{etype}]: {e}")
                        if etype not in self.ignored_error_types:
                            ignore = self._ask_from_thread(
                                _("import.write_err.title"),
                                _("import.write_err.msg", lineno=lineno, etype=etype, e=e),
                            )
                            if ignore:
                                self.ignored_error_types.add(etype)
                                self._imp_log(_("import.log.ignore_type", etype=etype))
                            else:
                                db.commit()
                                db.close()
                                self.after(0, lambda: (
                                    self.import_status_var.set(_("import.status.terminated")),
                                    self.import_btn.configure(state="normal"),
                                ))
                                return

                # 停止信号检测
                if self._stop_import.is_set():
                    db.rollback()
                    db.close()
                    im, sk, er = imported, skipped, errors
                    self.after(0, lambda a=im, b=sk, c=er: self._on_import_stopped(a, b, c))
                    return

                # 进度更新
                if idx % 500 == 0 or idx == total_data - 1:
                    pct = int((idx + 1) / total_data * 100)
                    im, sk, er = imported, skipped, errors
                    self.after(0, lambda p=pct, a=im, b=sk, c=er: self._update_imp_progress(p, a, b, c))

            # 剩余批次（rollback 模式）
            if error_mode == "rollback" and batch:
                try:
                    flush(batch)
                except Exception as e:
                    db.rollback()
                    db.close()
                    self._finish_import_error(_("import.err.write_fail_all", e=e))
                    return

            if error_mode == "skip":
                db.commit()

            db.close()
            msg = _("import.status.done", imported=imported, skipped=skipped, errors=errors)
            self.logger.info(msg)
            self.after(0, lambda: self._on_import_done(imported, skipped, errors, msg))

        except Exception as e:
            db.close()
            self._finish_import_error(str(e))

    def _stop_import_clicked(self):
        self._stop_import.set()
        self.stop_btn.configure(state="disabled")
        self.import_status_var.set(_("import.status.stopping"))

    def _ask_from_thread(self, title: str, msg: str) -> bool:
        """在后台线程中安全地弹出 Yes/No 对话框，阻塞直到用户回应。"""
        result = [False]
        evt = threading.Event()

        def ask():
            result[0] = messagebox.askyesno(title, msg)
            evt.set()

        self.after(0, ask)
        evt.wait()
        return result[0]

    def _on_import_stopped(self, imported, skipped, errors):
        self.imported_var.set(str(imported))
        self.skipped_var.set(str(skipped))
        self.errors_var.set(str(errors))
        self.import_status_var.set(_("import.status.stopped", n=imported))
        self.import_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        msg = _("import.stopped.msg", n=imported)
        self.logger.info(msg)
        self._imp_log(_("import.log.stopped", sep="="*60, msg=msg))
        messagebox.showwarning(_("import.stopped.title"), msg)

    def _update_imp_progress(self, pct, imported, skipped, errors):
        self.imp_pb["value"] = pct
        self.imported_var.set(str(imported))
        self.skipped_var.set(str(skipped))
        self.errors_var.set(str(errors))

    def _on_import_done(self, imported, skipped, errors, msg):
        self.imp_pb["value"] = 100
        self.imported_var.set(str(imported))
        self.skipped_var.set(str(skipped))
        self.errors_var.set(str(errors))
        self.import_status_var.set(msg)
        self.import_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self._imp_log(_("import.log.done", sep="="*60, msg=msg, log_path=self.log_path))
        messagebox.showinfo(_("import.done.title"), msg)

    def _finish_import_error(self, msg: str):
        self.logger.error(msg)
        self.after(0, lambda: (
            messagebox.showerror(_("import.fail.title"), msg),
            self.import_status_var.set(_("import.status.fail")),
            self.import_btn.configure(state="normal"),
            self.stop_btn.configure(state="disabled"),
        ))

    def _imp_log_clear(self):
        self.imp_log.configure(state=tk.NORMAL)
        self.imp_log.delete("1.0", tk.END)
        self.imp_log.configure(state=tk.DISABLED)

    def _imp_log(self, text: str):
        self.after(0, lambda t=text: self._imp_log_main(t))

    def _imp_log_main(self, text: str):
        self.imp_log.configure(state=tk.NORMAL)
        self.imp_log.insert(tk.END, text)
        self.imp_log.see(tk.END)
        self.imp_log.configure(state=tk.DISABLED)

    # ══════════════════════════════════════════════════════════════════════════
    # Tab 6 — 数据导出
    # ══════════════════════════════════════════════════════════════════════════
    def _build_export_tab(self):
        outer = ttk.Frame(self.nb)
        self.nb.add(outer, text=_("tab.export"))

        # 滚动容器
        canvas = tk.Canvas(outer, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        f = ttk.Frame(canvas)
        _wid = canvas.create_window((0, 0), window=f, anchor="nw")
        f.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_wid, width=e.width))
        # 鼠标滚轮仅在悬停时绑定，避免影响其他 tab
        canvas.bind("<Enter>", lambda e: canvas.bind_all(
            "<MouseWheel>", lambda ev: canvas.yview_scroll(int(-1 * ev.delta / 120), "units")))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # ── 数据来源 ──────────────────────────────────────────────────────────
        src_lf = ttk.LabelFrame(f, text=_("export.src.label"))
        src_lf.pack(fill=tk.X, padx=12, pady=(10, 4))

        src_row = ttk.Frame(src_lf)
        src_row.pack(fill=tk.X, padx=8, pady=6)
        ttk.Radiobutton(src_row, text=_("export.src.file"),
                        variable=self._export_source_var, value="file",
                        command=self._on_export_source_change).pack(side=tk.LEFT)
        ttk.Radiobutton(src_row, text=_("export.src.sql"),
                        variable=self._export_source_var, value="sql",
                        command=self._on_export_source_change).pack(side=tk.LEFT, padx=16)

        self._sql_input_frame = ttk.Frame(src_lf)
        self._sql_input_frame.pack(fill=tk.X, padx=8, pady=(0, 6))
        ttk.Label(self._sql_input_frame,
                  text=_("export.src.sql_hint")).pack(anchor=tk.W)
        self.export_sql_text = scrolledtext.ScrolledText(
            self._sql_input_frame, height=3, font=("Consolas", 9), wrap=tk.NONE)
        self.export_sql_text.pack(fill=tk.X, pady=(2, 0))
        self._sql_input_frame.pack_forget()

        # ── 导出配置 ──────────────────────────────────────────────────────────
        cfg_lf = ttk.LabelFrame(f, text=_("export.cfg.label"))
        cfg_lf.pack(fill=tk.X, padx=12, pady=4)

        row1 = ttk.Frame(cfg_lf)
        row1.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(row1, text=_("export.cfg.format")).pack(side=tk.LEFT)
        ttk.Radiobutton(row1, text="CSV", variable=self._export_format_var,
                        value="csv", command=self._on_export_format_change).pack(side=tk.LEFT, padx=4)
        ttk.Radiobutton(row1, text="Excel (.xlsx)", variable=self._export_format_var,
                        value="excel", command=self._on_export_format_change).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(row1, text=_("export.cfg.include_hdr"),
                        variable=self._export_include_header_var).pack(side=tk.LEFT, padx=20)

        # Excel 引擎选择行（仅 Excel 格式时可见）
        self._engine_row = ttk.Frame(cfg_lf)
        ttk.Label(self._engine_row, text=_("export.cfg.engine")).pack(side=tk.LEFT, padx=(0, 6))
        if HAS_XLSXWRITER:
            hint = _("export.cfg.engine_fast") if not HAS_OPENPYXL else _("export.cfg.engine_fast2")
            ttk.Radiobutton(self._engine_row, text=f"xlsxwriter {hint}",
                            variable=self._excel_engine_var,
                            value="xlsxwriter").pack(side=tk.LEFT)
        if HAS_OPENPYXL:
            hint = _("export.cfg.engine_compat") if HAS_XLSXWRITER else _("export.cfg.engine_default")
            ttk.Radiobutton(self._engine_row, text=f"openpyxl {hint}",
                            variable=self._excel_engine_var,
                            value="openpyxl").pack(side=tk.LEFT, padx=(8, 0))
        if not HAS_XLSXWRITER and not HAS_OPENPYXL:
            ttk.Label(self._engine_row, text=_("export.cfg.engine_none"),
                      foreground="red").pack(side=tk.LEFT)
        self._engine_row.pack_forget()  # 默认隐藏（CSV 模式）

        row2 = ttk.Frame(cfg_lf)
        row2.pack(fill=tk.X, padx=8, pady=(0, 6))
        ttk.Label(row2, text=_("export.cfg.outdir")).pack(side=tk.LEFT)
        self.export_dir_var = tk.StringVar(value=_load_ui_state().get("last_export_dir", APP_DIR))
        ttk.Entry(row2, textvariable=self.export_dir_var, width=42).pack(side=tk.LEFT, padx=4)
        ttk.Button(row2, text=_("export.cfg.browse"), command=self._browse_export_dir).pack(side=tk.LEFT)
        ttk.Label(row2, text=_("export.cfg.prefix")).pack(side=tk.LEFT, padx=(12, 2))
        self.export_prefix_var = tk.StringVar(value="export")
        ttk.Entry(row2, textvariable=self.export_prefix_var, width=14).pack(side=tk.LEFT)

        # ── 列配置 ────────────────────────────────────────────────────────────
        col_lf = ttk.LabelFrame(f, text=_("export.col.label"))
        col_lf.pack(fill=tk.X, padx=12, pady=4)

        col_ctrl = ttk.Frame(col_lf)
        col_ctrl.pack(fill=tk.X, padx=8, pady=6)
        ttk.Button(col_ctrl, text=_("export.col.load"),
                   command=self._load_export_columns).pack(side=tk.LEFT)
        self.export_col_info_var = tk.StringVar(value="")
        ttk.Label(col_ctrl, textvariable=self.export_col_info_var,
                  foreground="blue").pack(side=tk.LEFT, padx=12)
        ttk.Button(col_ctrl, text=_("export.col.select_all"),
                   command=lambda: self._set_all_export_cols("✓")).pack(side=tk.LEFT, padx=4)
        ttk.Button(col_ctrl, text=_("export.col.select_none"),
                   command=lambda: self._set_all_export_cols("☐")).pack(side=tk.LEFT, padx=2)

        col_tree_frm = ttk.Frame(col_lf)
        col_tree_frm.pack(fill=tk.X, padx=8, pady=(0, 6))
        self.col_tree = ttk.Treeview(col_tree_frm,
                                     columns=("include", "orig", "export"),
                                     show="headings", height=6, selectmode="browse")
        col_vsb = ttk.Scrollbar(col_tree_frm, orient=tk.VERTICAL, command=self.col_tree.yview)
        self.col_tree.configure(yscrollcommand=col_vsb.set)
        self.col_tree.heading("include", text=_("export.col.hdr_include"))
        self.col_tree.heading("orig",    text=_("export.col.hdr_orig"))
        self.col_tree.heading("export",  text=_("export.col.hdr_export"))
        self.col_tree.column("include", width=50, minwidth=50, stretch=False, anchor=tk.CENTER)
        self.col_tree.column("orig", width=200, stretch=True)
        self.col_tree.column("export", width=200, stretch=True)
        col_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.col_tree.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.col_tree.bind("<Button-1>", self._on_col_click)
        self.col_tree.bind("<Double-1>", self._on_col_double_click)

        # ── 导出方式 ──────────────────────────────────────────────────────────
        way_lf = ttk.LabelFrame(f, text=_("export.way.label"))
        way_lf.pack(fill=tk.X, padx=12, pady=4)

        way_row = ttk.Frame(way_lf)
        way_row.pack(fill=tk.X, padx=8, pady=6)
        ttk.Radiobutton(way_row, text=_("export.way.all"),
                        variable=self._export_split_var, value="all",
                        command=self._on_export_split_change).pack(side=tk.LEFT)
        ttk.Radiobutton(way_row, text=_("export.way.split"),
                        variable=self._export_split_var, value="split",
                        command=self._on_export_split_change).pack(side=tk.LEFT, padx=16)

        self._split_cfg_frame = ttk.Frame(way_lf)
        self._split_cfg_frame.pack(fill=tk.X, padx=8, pady=(0, 6))
        ttk.Label(self._split_cfg_frame, text=_("export.way.recs")).pack(side=tk.LEFT)
        ttk.Entry(self._split_cfg_frame, textvariable=self._export_records_per_file_var,
                  width=10).pack(side=tk.LEFT, padx=4)
        self._split_cfg_frame.pack_forget()

        # ── 执行区 ────────────────────────────────────────────────────────────
        btn_frm = ttk.Frame(f)
        btn_frm.pack(fill=tk.X, padx=12, pady=6)
        self.export_btn = ttk.Button(btn_frm, text=_("export.btn.start"), command=self._start_export)
        self.export_btn.pack(side=tk.LEFT, padx=4)
        self.export_stop_btn = ttk.Button(btn_frm, text=_("export.btn.stop"),
                                          command=self._stop_export_clicked,
                                          state="disabled")
        self.export_stop_btn.pack(side=tk.LEFT, padx=4)
        self.export_status_var = tk.StringVar(value="")
        ttk.Label(btn_frm, textvariable=self.export_status_var).pack(side=tk.LEFT, padx=10)

        self.export_pb = ttk.Progressbar(f, mode="determinate")
        self.export_pb.pack(fill=tk.X, padx=12, pady=4)

        ttk.Label(f, text=_("export.log.label")).pack(anchor=tk.W, padx=12, pady=(4, 2))
        self.export_log = scrolledtext.ScrolledText(
            f, height=7, state=tk.DISABLED, font=("Consolas", 9))
        self.export_log.pack(fill=tk.BOTH, padx=12, pady=(0, 10))

    # ── 导出 Tab 事件 ─────────────────────────────────────────────────────────

    def _on_export_source_change(self):
        if self._export_source_var.get() == "sql":
            self._sql_input_frame.pack(fill=tk.X, padx=8, pady=(0, 6))
        else:
            self._sql_input_frame.pack_forget()

    def _on_export_format_change(self):
        if self._export_format_var.get() == "excel":
            self._engine_row.pack(fill=tk.X, padx=8, pady=(0, 6))
        else:
            self._engine_row.pack_forget()

    def _on_export_split_change(self):
        if self._export_split_var.get() == "split":
            self._split_cfg_frame.pack(fill=tk.X, padx=8, pady=(0, 6))
        else:
            self._split_cfg_frame.pack_forget()

    def _browse_export_dir(self):
        d = filedialog.askdirectory(title=_("export.cfg.outdir").rstrip(":"))
        if d:
            self.export_dir_var.set(d)
            _save_ui_state({"last_export_dir": d})

    def _load_export_columns(self):
        source = self._export_source_var.get()
        if source == "file":
            vr = self.validation_result
            if not vr:
                messagebox.showwarning(_("export.warn.title"), _("export.warn.no_validate"))
                return
            columns    = vr["columns"]
            total_data = vr["data_count"]
            inv_count  = len(vr["invalid"])
            valid_count = total_data - inv_count
            self.export_col_info_var.set(
                _("export.col_info_valid", valid=valid_count, total=total_data, inv=inv_count))
        else:
            sql = self.export_sql_text.get("1.0", tk.END).strip()
            if not sql:
                messagebox.showwarning(_("export.warn.title"), _("export.warn.no_sql_text"))
                return
            try:
                db = DBConnection(self._get_ui_db_cfg())
                db.connect()
                cur = db.cursor()
                # 只取列名，不拉取数据（防止大结果集撑爆内存）
                cur.execute(sql)
                columns = [desc[0] for desc in cur.description]
                cur.close()
                db.close()
                self._sql_export_sql = sql
                self._sql_export_columns = columns
                self.export_col_info_var.set(_("export.col_info_sql", cols=len(columns)))
            except Exception as e:
                messagebox.showerror(_("export.err.sql"), str(e))
                return

        for item_id in self.col_tree.get_children():
            self.col_tree.delete(item_id)
        for col in columns:
            self.col_tree.insert("", tk.END, values=("✓", col, col))

    def _set_all_export_cols(self, state: str):
        for item_id in self.col_tree.get_children():
            self.col_tree.set(item_id, "include", state)

    def _on_col_click(self, event):
        if self.col_tree.identify_region(event.x, event.y) != "cell":
            return
        col = self.col_tree.identify_column(event.x)
        row = self.col_tree.identify_row(event.y)
        if not row or col != "#1":
            return
        cur = self.col_tree.set(row, "include")
        self.col_tree.set(row, "include", "☐" if cur == "✓" else "✓")

    def _on_col_double_click(self, event):
        col = self.col_tree.identify_column(event.x)
        row = self.col_tree.identify_row(event.y)
        if not row or col != "#3":
            return
        bbox = self.col_tree.bbox(row, "export")
        if not bbox:
            return
        x, y, width, height = bbox
        var = tk.StringVar(value=self.col_tree.set(row, "export"))
        entry = ttk.Entry(self.col_tree, textvariable=var)
        entry.place(x=x, y=y, width=width, height=height)
        entry.select_range(0, tk.END)
        entry.focus_set()

        def _commit(e=None):
            new = var.get().strip()
            self.col_tree.set(row, "export", new if new else self.col_tree.set(row, "orig"))
            entry.destroy()

        entry.bind("<Return>", _commit)
        entry.bind("<FocusOut>", _commit)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def _get_export_col_config(self):
        """Returns [(orig_name, export_name), ...] for selected columns."""
        result = []
        for item_id in self.col_tree.get_children():
            if self.col_tree.set(item_id, "include") == "✓":
                orig = self.col_tree.set(item_id, "orig")
                exp = self.col_tree.set(item_id, "export") or orig
                result.append((orig, exp))
        return result

    def _start_export(self):
        col_config = self._get_export_col_config()
        if not col_config:
            messagebox.showwarning(_("export.warn.title"), _("export.warn.no_col"))
            return

        output_dir = self.export_dir_var.get().strip()
        if not output_dir or not os.path.isdir(output_dir):
            messagebox.showwarning(_("export.warn.title"), _("export.warn.no_dir"))
            return

        prefix = self.export_prefix_var.get().strip() or "export"
        fmt = self._export_format_var.get()
        include_header = self._export_include_header_var.get()
        split = self._export_split_var.get() == "split"
        recs_per_file = 10000
        if split:
            try:
                recs_per_file = max(1, int(self._export_records_per_file_var.get() or 10000))
            except ValueError:
                recs_per_file = 10000

        engine = self._excel_engine_var.get()
        if fmt == "excel":
            if engine == "xlsxwriter" and not HAS_XLSXWRITER:
                messagebox.showerror(_("export.err.title"), _("export.err.no_xlsxwriter"))
                return
            if engine == "openpyxl" and not HAS_OPENPYXL:
                messagebox.showerror(_("export.err.title"), _("export.err.no_openpyxl"))
                return
            if not engine:
                messagebox.showerror(_("export.err.title"), _("export.err.no_engine"))
                return

        source = self._export_source_var.get()
        if source == "file" and not self.validation_result:
            messagebox.showwarning(_("export.warn.title"), _("export.warn.no_validate"))
            return
        if source == "sql" and not hasattr(self, "_sql_export_sql"):
            messagebox.showwarning(_("export.warn.title"), _("export.warn.no_sql"))
            return

        self._stop_export.clear()
        self.export_btn.configure(state="disabled")
        self.export_stop_btn.configure(state="normal")
        self.export_pb["value"] = 0
        self._export_log_clear()
        self.export_status_var.set(_("export.status.preparing"))

        threading.Thread(
            target=self._do_export,
            args=(col_config, output_dir, prefix, fmt, include_header, split, recs_per_file, engine),
            daemon=True,
        ).start()

    def _stop_export_clicked(self):
        self._stop_export.set()
        self.export_stop_btn.configure(state="disabled")
        self.export_status_var.set(_("export.status.stopping"))

    def _do_export(self, col_config, output_dir, prefix, fmt, include_header, split, recs_per_file, engine="openpyxl"):
        import itertools, re as _re, time as _time
        FETCH_SIZE = 10_000  # SQL 模式每批拉取行数

        # ── 导出专用日志文件 ──────────────────────────────────────────────────
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(LOG_DIR, exist_ok=True)
        log_fname  = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_path   = os.path.join(LOG_DIR, log_fname)
        exp_logger = make_logger(log_path)

        def _log(msg: str):
            exp_logger.info(msg)
            self._export_log_write(msg + "\n")

        # ── Excel 非法字符清洗（优化：只在行内有控制字符时才做替换）─────────
        _ILLEGAL_XML   = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
        _ILLEGAL_QUICK = _re.compile(r'[\x00-\x1f]')   # 更宽松的预筛：ASCII 控制字符

        def _sanitize_row(row):
            """
            对一行字段做 XML 非法字符清洗。
            先用快速 pattern 扫描整行拼接字符串；若无命中直接返回原行（零分配）。
            """
            # 只对字符串字段拼接做快速预扫
            concat = "".join(v for v in row if isinstance(v, str))
            if not concat or not _ILLEGAL_QUICK.search(concat):
                return row   # 绝大多数行走这条快速路径
            # 确有控制字符，精确替换
            return [_ILLEGAL_XML.sub(' ', v) if isinstance(v, str) else v for v in row]

        try:
            t0     = _time.time()
            source = self._export_source_var.get()
            export_headers = [exp for _, exp in col_config]
            db = None

            _log(_("export.log.start"))
            _log(_("export.log.source", src=source))
            _log(_("export.log.fmt", fmt=fmt, engine=engine if fmt != 'csv' else 'N/A'))
            _log(_("export.log.outdir", path=output_dir))
            _log(_("export.log.prefix", prefix=prefix))
            split_mode = _("export.log.split_yes", n=recs_per_file) if split else _("export.log.split_no")
            _log(_("export.log.split", mode=split_mode))
            _log(_("export.log.cols", cols=export_headers))

            # ── 准备流式数据源 ──────────────────────────────────────────────────
            if source == "file":
                vr = self.validation_result
                col_names = vr["columns"]
                col_idx    = {c: i for i, c in enumerate(col_names)}
                invalid_ln = {item[0] for item in vr["invalid"]}
                total      = vr["data_count"] - len(vr["invalid"])

                def _file_gen():
                    # 流式重读文件，不依赖内存中的 all_rows
                    for lineno, row in self._stream_data_rows(vr):
                        if lineno not in invalid_ln:
                            yield [row[col_idx[orig]] if orig in col_idx and col_idx[orig] < len(row) else ""
                                   for orig, _ in col_config]
                data_gen = _file_gen()

            else:
                sql = self._sql_export_sql
                col_names = self._sql_export_columns
                col_idx = {c: i for i, c in enumerate(col_names)}
                is_csv = (fmt == "csv")

                db = DBConnection(self._get_ui_db_cfg())
                db.connect()
                # 尝试提前获取总行数（COUNT 失败时继续，仅影响进度百分比）
                total = None
                try:
                    cc = db.cursor()
                    cc.execute(f"SELECT COUNT(*) FROM ({sql}) _c")
                    total = cc.fetchone()[0]
                    cc.close()
                except Exception:
                    pass

                cur = db.cursor()
                cur.execute(sql)

                def _sql_gen():
                    while True:
                        batch = cur.fetchmany(FETCH_SIZE)
                        if not batch:
                            break
                        for row in batch:
                            yield [
                                (str(row[col_idx[orig]]) if is_csv else row[col_idx[orig]])
                                if orig in col_idx and row[col_idx[orig]] is not None else ""
                                for orig, _ in col_config
                            ]
                data_gen = _sql_gen()

            # ── 导出（流式写入，不再全量缓存）──────────────────────────────────
            processed = 0
            file_count = 0
            file_idx = 0

            def _update_progress():
                if total:
                    pct = min(processed / total * 100, 99)
                    self.after(0, lambda p=pct: self.export_pb.configure(value=p))
                elapsed = _time.time() - t0
                speed   = int(processed / elapsed) if elapsed > 0 else 0
                if total:
                    msg = _("export.progress", done=processed, total=total, speed=speed)
                else:
                    msg = _("export.progress_nototal", done=processed, speed=speed)
                self.after(0, lambda m=msg: self.export_status_var.set(m))
                exp_logger.info(msg)

            def _write_chunk(rows_chunk):
                """将一批行写入单个文件，返回文件路径和实际写入行数。"""
                suffix = f"_{file_idx + 1:03d}" if split else ""
                t_chunk = _time.time()
                if fmt == "csv":
                    fname = f"{prefix}{suffix}.csv"
                    fpath = os.path.join(output_dir, fname)
                    _log(_("export.log.file_start", path=fpath))
                    with open(fpath, "w", encoding="utf-8-sig", newline="") as fh:
                        writer = csv.writer(fh)
                        if include_header:
                            writer.writerow(export_headers)
                        cnt = 0
                        for row in rows_chunk:
                            writer.writerow(row)
                            cnt += 1
                    elapsed_c = _time.time() - t_chunk
                    _log(_("export.log.file_done", path=fpath, cnt=cnt, elapsed=elapsed_c,
                            speed=int(cnt/elapsed_c) if elapsed_c > 0 else 0))
                    return fpath, cnt
                elif engine == "xlsxwriter":
                    import xlsxwriter
                    fname = f"{prefix}{suffix}.xlsx"
                    fpath = os.path.join(output_dir, fname)
                    _log(_("export.log.file_start", path=fpath))
                    wb = xlsxwriter.Workbook(fpath, {'constant_memory': True, 'use_zip64': True})
                    ws = wb.add_worksheet("Sheet1")
                    row_idx = 0
                    if include_header:
                        ws.write_row(row_idx, 0, export_headers)
                        row_idx += 1
                    cnt = 0
                    t_write = _time.time()
                    for row in rows_chunk:
                        ws.write_row(row_idx, 0, _sanitize_row(row))
                        row_idx += 1
                        cnt += 1
                    t_after_write = _time.time()
                    _log(_("export.log.write_row", engine="xlsxwriter", cnt=cnt,
                            elapsed=t_after_write - t_write,
                            speed=int(cnt/(t_after_write-t_write)) if t_after_write > t_write else 0))
                    wb.close()
                    t_after_close = _time.time()
                    _log(_("export.log.close_zip", engine="xlsxwriter",
                            elapsed=t_after_close - t_after_write))
                    elapsed_c = t_after_close - t_chunk
                    _log(_("export.log.file_done", path=fpath, cnt=cnt, elapsed=elapsed_c,
                            speed=int(cnt/elapsed_c) if elapsed_c > 0 else 0))
                    return fpath, cnt
                else:  # openpyxl
                    import openpyxl
                    t_open = _time.time()
                    wb = openpyxl.Workbook(write_only=True)
                    ws = wb.create_sheet("Sheet1")
                    fname = f"{prefix}{suffix}.xlsx"
                    fpath = os.path.join(output_dir, fname)
                    _log(_("export.log.file_start", path=fpath))
                    _log(_("export.log.wb_init", engine="openpyxl",
                            elapsed=_time.time() - t_open))
                    if include_header:
                        ws.append(export_headers)
                    cnt = 0
                    t_append = _time.time()
                    for row in rows_chunk:
                        ws.append(_sanitize_row(row))
                        cnt += 1
                    t_after_append = _time.time()
                    _log(_("export.log.append", engine="openpyxl", cnt=cnt,
                            elapsed=t_after_append - t_append,
                            speed=int(cnt/(t_after_append-t_append)) if t_after_append > t_append else 0))
                    wb.save(fpath)
                    t_after_save = _time.time()
                    _log(_("export.log.save_zip", engine="openpyxl",
                            elapsed=t_after_save - t_after_append))
                    elapsed_c = t_after_save - t_chunk
                    _log(_("export.log.file_done", path=fpath, cnt=cnt, elapsed=elapsed_c,
                            speed=int(cnt/elapsed_c) if elapsed_c > 0 else 0))
                    return fpath, cnt

            if split:
                # 分片：每次从生成器取 recs_per_file 行，逐文件写入
                while not self._stop_export.is_set():
                    chunk_rows = list(itertools.islice(data_gen, recs_per_file))
                    if not chunk_rows:
                        break

                    def _tracked(rows):
                        nonlocal processed
                        for row in rows:
                            if self._stop_export.is_set():
                                return
                            yield row
                            processed += 1
                            if processed % 5000 == 0:
                                _update_progress()

                    fpath, cnt = _write_chunk(_tracked(chunk_rows))
                    file_count += 1
                    file_idx += 1
                    self._export_log_write(_("export.log.written", path=fpath, cnt=cnt))
            else:
                # stream all rows to single file
                def _tracked_all():
                    nonlocal processed
                    for row in data_gen:
                        if self._stop_export.is_set():
                            return
                        yield row
                        processed += 1
                        if processed % 5000 == 0:
                            _update_progress()

                fpath, cnt = _write_chunk(_tracked_all())
                file_count = 1
                self._export_log_write(_("export.log.written", path=fpath, cnt=cnt))

            if db:
                db.close()

            if self._stop_export.is_set():
                msg = _("export.stopped", done=processed, files=file_count)
            else:
                total_val = total if total else processed
                msg = _("export.done", total=f"{total_val:,}", files=file_count)

            self.after(0, lambda m=msg: (
                self.export_pb.configure(value=100),
                self.export_status_var.set(m),
                self.export_btn.configure(state="normal"),
                self.export_stop_btn.configure(state="disabled"),
            ))

        except Exception as e:
            import traceback
            err = str(e)
            tb = traceback.format_exc()
            self._export_log_write(_("export.err.exception", tb=tb))
            self.logger.error(f"export error: {tb}")
            self.after(0, lambda: (
                self.export_status_var.set(_("export.err.msg", err=err)),
                self.export_btn.configure(state="normal"),
                self.export_stop_btn.configure(state="disabled"),
            ))

    def _export_log_clear(self):
        self.export_log.configure(state=tk.NORMAL)
        self.export_log.delete("1.0", tk.END)
        self.export_log.configure(state=tk.DISABLED)

    def _export_log_write(self, text: str):
        self.after(0, lambda t=text: self._export_log_write_main(t))

    def _export_log_write_main(self, text: str):
        self.export_log.configure(state=tk.NORMAL)
        self.export_log.insert(tk.END, text)
        self.export_log.see(tk.END)
        self.export_log.configure(state=tk.DISABLED)

    # ── Info Tab ──────────────────────────────────────────────────────────────

    def _build_info_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text=_("tab.info"))

        # ── 语言设置 ──────────────────────────────────────────────────────────
        lang_lf = ttk.LabelFrame(f, text=_("info.lang.label"))
        lang_lf.pack(fill=tk.X, padx=20, pady=(20, 8))

        lang_row = ttk.Frame(lang_lf)
        lang_row.pack(fill=tk.X, padx=12, pady=10)

        self._lang_var = tk.StringVar(value=current_lang())
        for lang_code in available_langs():
            label = LANG_LABELS.get(lang_code, lang_code)
            ttk.Radiobutton(
                lang_row, text=label,
                variable=self._lang_var, value=lang_code,
                command=self._on_lang_change,
            ).pack(side=tk.LEFT, padx=10)

        self._lang_hint_var = tk.StringVar(value="")
        ttk.Label(lang_lf, textvariable=self._lang_hint_var,
                  foreground="gray").pack(anchor=tk.W, padx=12, pady=(0, 8))

        # ── 关于 ──────────────────────────────────────────────────────────────
        about_lf = ttk.LabelFrame(f, text=_("info.tab.title"))
        about_lf.pack(fill=tk.X, padx=20, pady=8)

        info_lines = [
            _("info.app.name"),
            _("info.app.desc"),
            _("info.app.support"),
            "",
            _("info.app.license"),
        ]
        for line in info_lines:
            ttk.Label(about_lf, text=line).pack(anchor=tk.W, padx=16, pady=2)
        ttk.Frame(about_lf).pack(pady=4)  # bottom padding

    def _on_lang_change(self):
        lang = self._lang_var.get()
        set_lang(lang)
        _save_ui_state({"lang": lang})
        self._lang_hint_var.set(_("info.lang.restart_hint"))


# ─────────────────────────────────────────────────────────────────────────────
# 辅助
# ─────────────────────────────────────────────────────────────────────────────

def delimiter_join(row: list, delimiter: str) -> str:
    return delimiter.join(str(v) for v in row)


# ─────────────────────────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = CSVImporterApp()
    app.mainloop()
